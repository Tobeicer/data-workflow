from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from product_profile import analyze_price, parse_moq_number, parse_unit_text


PRODUCT_FIELDS = {
    "source_platform": "来源平台",
    "product_id": "商品ID",
    "product_url": "商品链接",
    "title": "商品标题",
    "observed_at": "采集时间",
    "manufacturer_id": "厂家ID",
    "manufacturer_member_id": "厂家memberId",
    "manufacturer_name": "厂家名称",
    "manufacturer_relation_type": "商品厂家关系",
    "manufacturer_relation_status": "关联状态",
    "youyiquan_category_candidate": "游艺圈分类候选",
    "product_category": "产品类别",
    "brand": "品牌",
    "model": "型号",
    "item_number": "货号",
    "function": "功能",
    "material": "材质",
    "origin": "产地",
    "applicable_people": "适用人数",
    "applicable_age": "适用年龄",
    "applicable_scenarios": "适用场景",
    "technical_type": "技术类型",
    "color": "颜色",
    "ip_authorized": "是否IP授权",
    "ccc_configuration_category": "3C配置类别",
    "ccc_certificate_code": "商品3C认证码",
    "packaging": "包装",
    "patent_type": "专利类型",
    "price_min": "最低价(元)",
    "price_max": "最高价(元)",
    "currency": "货币",
    "price_status": "价格状态",
    "price_missing_reason": "价格缺失原因",
    "minimum_order_quantity": "起订量",
    "sales_unit": "销售单位",
    "available_stock": "可售库存",
    "delivery_commitment": "发货承诺",
    "quality_report_number": "质检报告编号",
    "pack_length_cm": "包装长(cm)",
    "pack_width_cm": "包装宽(cm)",
    "pack_height_cm": "包装高(cm)",
    "pack_volume_cm3": "包装体积(cm³)",
    "pack_weight_g": "包装重量(g)",
    "pack_specs_json": "包装明细(JSON)",
    "main_image_url": "商品主图",
    "image_urls": "商品图片",
    "video_url": "商品视频",
    "detail_content_url": "详情内容链接",
    "detail_images_json": "详情图片(JSON)",
    "service_guarantees": "服务保障",
    "raw_sku_count": "原始SKU数量",
    "related_product_count": "相关商品数",
    "related_products_json": "相关商品(JSON)",
    "other_attributes": "其他商品属性",
}


DELIVERY_SKU_FIELDS = {
    "product_id": "商品ID",
    "sku_name": "SKU名称",
    "sku_price": "SKU价格(元)",
    "sku_price_text": "SKU价格原文",
    "stock_quantity": "SKU库存数量",
    "sku_image_url": "SKU图片",
    "collected_at": "采集时间",
}


MANUFACTURER_FIELDS = {
    "source_platform": "来源平台",
    "manufacturer_id": "厂家ID",
    "member_id": "1688 memberId",
    "company_id": "1688公司ID",
    "manufacturer_name": "厂家名称",
    "related_product_ids": "关联商品ID",
    "product_relation_type": "商品厂家关系",
    "shop_url": "店铺链接",
    "factory_archive_url": "工厂档案链接",
    "subject_qualification_url": "主体资质入口",
    "business_info_url": "工商详情链接",
    "unified_social_credit_code": "统一社会信用代码",
    "legal_representative": "法定代表人",
    "registered_capital": "注册资本",
    "established_date": "成立日期",
    "company_type": "企业类型",
    "registration_authority": "登记机关",
    "business_term": "营业期限",
    "registered_address": "注册地址",
    "business_scope": "经营范围",
    "contact_person": "联系人",
    "telephone": "联系电话",
    "mobile": "手机",
    "contact_address": "联系地址",
    "main_category": "主营类目",
    "production_service": "主营产品/服务",
    "company_summary": "公司简介",
    "brands": "自有品牌",
    "factory_address": "工厂地址",
    "factory_area_sqm": "工厂面积（㎡）",
    "factory_area_authenticated": "工厂面积已认证",
    "employee_total_range": "员工规模",
    "production_line_count": "生产线数量",
    "production_equipment_count": "生产设备数量",
    "annual_transaction_amount": "年交易额",
    "monthly_output_value": "月产值",
    "processing_methods": "加工方式",
    "custom_minimum_order": "定制起订量",
    "vat_invoice_available": "可开增值税发票",
    "qualification_tags": "工厂能力/认证标签",
    "certificate_count": "资质证书数量",
    "certificates": "资质证书明细",
    "factory_auth_provider": "深度认证机构",
    "factory_auth_report_number": "认证报告编号",
    "patent_count": "专利数量",
    "patents": "专利明细",
    "factory_medal": "工厂牌级",
    "returning_customer_rate": "回头率",
    "service_response_rate": "服务响应率",
    "on_time_fulfillment_rate": "准时履约率",
    "factory_vr_url": "工厂VR展厅",
    "factory_images": "工厂图片",
    "factory_videos": "工厂视频",
    "observed_at": "采集时间",
}


def repair_mojibake(text: str) -> str:
    candidates = [text]
    for encoding in ("utf-8", "gb18030"):
        try:
            candidates.append(text.encode("latin-1").decode(encoding))
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass

    def score(candidate: str) -> int:
        chinese = sum("\u4e00" <= char <= "\u9fff" for char in candidate)
        latin_noise = sum("\u00c0" <= char <= "\u00ff" for char in candidate)
        return chinese * 6 - latin_noise * 2 - candidate.count("�") * 20

    return max(candidates, key=score)


def fmt_price(value: Any) -> str:
    """价格值 -> 2 位小数（元/分规范）。原始精度保留在 sku_price_text/price_text 证据。"""
    text = clean(value)
    if not text:
        return ""
    try:
        num = float(text)
    except ValueError:
        return ""
    if num == int(num):
        return str(int(num))
    return "%.2f" % num


def clean(value: Any) -> str:
    if value is None:
        return ""
    return repair_mojibake(str(value).strip())


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def nested(value: Any, *keys: str, default: Any = "") -> Any:
    current = value
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    return default if current is None else current


def unique(values: list[Any]) -> list[Any]:
    result: list[Any] = []
    seen: set[str] = set()
    for value in values:
        marker = json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list)) else clean(value)
        if not marker or marker in seen:
            continue
        seen.add(marker)
        result.append(value)
    return result


def select_snapshot(asset: dict, snapshot_type: str) -> dict:
    for item in asset.get("factory_snapshots") or []:
        if isinstance(item, dict) and item.get("snapshot_type") == snapshot_type:
            return item
    return {}


def asset_rank(asset: dict) -> tuple[int, str]:
    has_factory_archive = bool(select_snapshot(asset, "factory_archive_page"))
    return (1 if has_factory_archive else 0, clean(asset.get("collected_at")))


def product_rank(product: dict) -> tuple[str, int]:
    title = clean(product.get("title"))
    chinese_count = sum("\u4e00" <= char <= "\u9fff" for char in title)
    return (clean(product.get("collected_at")), chinese_count)


def manufacturer_name(asset: dict, fallback: str) -> str:
    return clean(nested(asset, "company", "company_name")) or clean(fallback)


def product_record(product: dict, asset: dict) -> dict:
    attributes = product.get("attributes") or {}
    source_fields = product.get("source_fields") or {}
    root_data = nested(source_fields, "Root", "fields", "dataJson", default={})
    gallery = nested(source_fields, "gallery", "fields", default={})
    description = nested(source_fields, "description", "fields", default={})
    services = nested(source_fields, "mainServices", "fields", default={})
    shipping = nested(source_fields, "shippingServices", "fields", default={})
    title_fields = nested(source_fields, "productTitle", "fields", default={})
    order_param = nested(root_data, "orderParamModel", "orderParam", default={})
    sku_map = nested(root_data, "skuModel", "skuInfoMap", default={})
    if not isinstance(sku_map, dict):
        sku_map = {}
    stock_values = [item.get("canBookCount") for item in sku_map.values() if isinstance(item, dict) and isinstance(item.get("canBookCount"), (int, float))]
    images = unique([clean(item) for item in (product.get("image_urls") or gallery.get("mainImage") or gallery.get("offerImgList") or [])])[:12]
    video = product.get("video") or gallery.get("video") or {}
    guarantees = unique(
        [
            clean(item.get("buyerDescription") or item.get("serviceName") or item.get("description"))
            for item in services.get("guaranteeList") or []
            if isinstance(item, dict)
        ]
    )
    member_id = clean(product.get("member_id"))
    factory = select_snapshot(asset, "factory_archive_page")
    relation_type = "factory_supplier_of" if factory else "source_supplier_of"
    relation_status = "source_factory_linked" if factory else "source_supplier_linked"
    name = manufacturer_name(asset, product.get("supplier_name") or product.get("shop_name") or "")
    mapped_attribute_keys = {
        "产品类别", "类型", "品牌", "型号", "货号", "功能", "材质", "产地", "适用人数",
        "适用年龄段", "适用年龄", "适用场景", "技术类型", "VR系统", "系统", "是否支持一件代发",
        "是否跨境出口专供货源", "出口认证", "质检报告编号", "售后服务",
        "颜色", "是否IP授权", "3C配置类别", "商品3C认证码", "屏幕类型", "屏幕尺寸",
        "分辨率", "接口", "亮度", "对比度", "包装", "用途", "专利类型",
        "主要下游平台", "主要销售地区", "有可授权的自有品牌",
    }
    other_attributes = {
        clean(key): clean(value)
        for key, value in attributes.items()
        if clean(key) not in mapped_attribute_keys and clean(value)
    }
    pack_specs = product.get("pack_specs") or []
    first_pack = pack_specs[0] if isinstance(pack_specs, list) and pack_specs else {}
    if not isinstance(first_pack, dict):
        first_pack = {}
    detail_images = product.get("detail_images") or []
    if not isinstance(detail_images, list):
        detail_images = []
    related_products = product.get("related_products") or []
    if not isinstance(related_products, list):
        related_products = []
    price_analysis_data = product.get("price_analysis") or {}
    if not isinstance(price_analysis_data, dict):
        price_analysis_data = {}
    sku_price_values = [
        clean(item.get("sku_price")) for item in (product.get("skus") or []) if isinstance(item, dict)
    ]
    sku_price_values = [v for v in sku_price_values if v]
    pa = price_analysis_data
    if not pa.get("price_status"):
        pa = analyze_price(clean(product.get("price_text")), sku_price_values)
    raw_color = clean(attributes.get("颜色"))
    color_value = (
        raw_color
        if raw_color
        and len(raw_color) <= 12
        and not any(ch in raw_color for ch in ",，、/【】")
        else ""
    )
    return {
        "source_platform": "1688",
        "product_id": clean(product.get("offer_id")),
        "product_url": clean(product.get("product_url")),
        "title": clean(product.get("title")),
        "manufacturer_id": f"1688:manufacturer:{member_id}",
        "manufacturer_member_id": member_id,
        "manufacturer_name": name,
        "manufacturer_relation_type": relation_type,
        "manufacturer_relation_status": relation_status,
        "youyiquan_category_candidate": clean(product.get("validation_category")),
        "product_category": clean(attributes.get("产品类别") or attributes.get("类型")),
        "brand": clean(attributes.get("品牌")),
        "model": clean(attributes.get("型号")),
        "item_number": clean(attributes.get("货号")),
        "function": clean(attributes.get("功能")),
        "material": clean(attributes.get("材质")),
        "origin": clean(attributes.get("产地")),
        "applicable_people": clean(attributes.get("适用人数")),
        "applicable_age": clean(attributes.get("适用年龄段") or attributes.get("适用年龄")),
        "applicable_scenarios": clean(attributes.get("适用场景")),
        "technical_type": clean(attributes.get("技术类型")),
        "color": color_value,
        "ip_authorized": clean(attributes.get("是否IP授权")),
        "ccc_configuration_category": clean(attributes.get("3C配置类别")),
        "ccc_certificate_code": clean(attributes.get("商品3C认证码")),
        "packaging": clean(attributes.get("包装")),
        "patent_type": clean(attributes.get("专利类型")),
        "price_min": clean(pa.get("price_min")),
        "price_max": clean(pa.get("price_max")),
        "currency": clean(pa.get("currency")) or "CNY",
        "price_status": clean(pa.get("price_status")),
        "price_missing_reason": clean(pa.get("price_missing_reason")),
        "minimum_order_quantity": parse_moq_number(product.get("minimum_order_quantity"))
        or parse_moq_number(order_param.get("beginNum")),
        "sales_unit": parse_unit_text(product.get("sales_unit"), product.get("moq_text"))
        or parse_unit_text(
            clean(title_fields.get("unit") or shipping.get("unit") or nested(root_data, "tempModel", "offerUnit"))
        ),
        "available_stock": (
            clean(product.get("available_stock"))
            or (int(sum(stock_values)) if stock_values else "")
        ),
        "delivery_commitment": clean(product.get("delivery_commitment"))
        or clean(shipping.get("deliveryLimitText")),
        "quality_report_number": clean(attributes.get("质检报告编号")),
        "other_attributes": other_attributes,
        "pack_length_cm": clean(first_pack.get("length_cm")),
        "pack_width_cm": clean(first_pack.get("width_cm")),
        "pack_height_cm": clean(first_pack.get("height_cm")),
        "pack_volume_cm3": clean(first_pack.get("volume_cm3")),
        "pack_weight_g": clean(first_pack.get("weight_g")),
        "pack_specs_json": json.dumps(pack_specs, ensure_ascii=False) if pack_specs else "",
        "detail_images_json": json.dumps(detail_images, ensure_ascii=False) if detail_images else "",
        "main_image_url": clean(product.get("main_image_url")) or (images[0] if images else ""),
        "image_urls": images,
        "video_url": clean(video.get("video_url") or video.get("videoUrl")),
        "detail_content_url": clean(product.get("detail_content_url") or description.get("detailUrl")),
        "service_guarantees": guarantees,
        "raw_sku_count": product.get("sku_count", ""),
        "related_product_count": len(related_products),
        "related_products_json": json.dumps(related_products[:20], ensure_ascii=False) if related_products else "",
        "observed_at": clean(product.get("collected_at")),
    }


def manufacturer_record(asset: dict, related_products: list[dict], fallback_name: str) -> dict:
    company = asset.get("company") or {}
    profile = asset.get("company_profile") or {}
    contacts = asset.get("contacts") or {}
    factory = select_snapshot(asset, "factory_archive_page")
    credit = select_snapshot(asset, "credit_detail_page")
    member_id = clean(company.get("member_id"))
    patents = asset.get("patent_details") or {}
    certificate_details = asset.get("certificate_details") or {}
    patent_items = [
        {"专利名称": clean(item.get("patent_name")), "专利号": clean(item.get("patent_number"))}
        for item in patents.get("items") or []
        if isinstance(item, dict)
    ]
    certificate_items = [
        {
            "证书名称": clean(item.get("certificate_name")),
            "证书链接": clean(item.get("certificate_url")),
        }
        for item in certificate_details.get("items") or []
        if isinstance(item, dict) and clean(item.get("certificate_name"))
    ]
    tags = unique([clean(item) for item in factory.get("factory_qualification_tags") or []])
    foreign_trade = clean(factory.get("foreign_trade_orders"))
    has_support_tag = "支持外贸订单" in tags
    if foreign_trade and has_support_tag and "不支持" in foreign_trade:
        foreign_trade_status = "conflict_review_required"
        foreign_trade_display = f"结构化字段：{foreign_trade}；能力标签：支持外贸订单"
    else:
        foreign_trade_status = "observed" if foreign_trade or has_support_tag else "not_observed"
        foreign_trade_display = foreign_trade or ("支持外贸订单" if has_support_tag else "")
    legal_keys = (
        "unified_social_credit_code", "registration_number", "legal_representative",
        "registered_capital_text", "established_date", "company_type", "registration_authority",
        "business_term", "registered_address", "business_scope",
    )
    legal_count = sum(bool(clean(company.get(key))) and clean(company.get(key)) != "暂无" for key in legal_keys)
    factory_url = clean(factory.get("source_url"))
    credit_url = clean(credit.get("source_url"))
    relation_type = "factory_supplier_of" if factory else "source_supplier_of"
    images = unique([item for item in factory.get("factory_images") or [] if isinstance(item, dict)])
    videos = unique([item for item in factory.get("factory_videos") or [] if isinstance(item, dict)])
    area = factory.get("factory_area_sqm", "")
    if area != "":
        area_path = "factory_archive_page.factory_area_sqm（API来源：factoryAreaData.relaDeepFactoryControlAcreage 或 fcProcessData.tagList[acreage]）"
    else:
        area_path = ""
    established_date = clean(company.get("established_date"))
    if not established_date:
        raw_established = clean(factory.get("established_time"))
        parts = re.findall(r"\d+", raw_established)
        if len(parts) >= 3:
            established_date = f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    factory_core_values = (
        area,
        factory.get("employee_count", ""),
        factory.get("production_line_count", ""),
        clean(factory.get("annual_transaction_amount_text")),
        clean(factory.get("factory_address")),
        clean(factory.get("factory_profile")),
        factory.get("brands") or [],
    )
    factory_is_sparse = bool(factory) and not any(value not in ("", None, []) for value in factory_core_values)
    if factory_is_sparse:
        quality_status = "factory_archive_sparse_source"
        quality_note = clean(factory.get("missing_reason")) or "工厂档案已访问；厂家未公开工厂面积、员工数、生产线、年交易额等资料，空值表示来源未披露。"
    elif factory:
        quality_status = "factory_archive_observed"
        quality_note = "工厂档案已采集；页面未展示的字段保持空值，表示厂家未上传或来源未披露。"
    else:
        quality_status = "company_pages_only"
        quality_note = "本批次仅有店铺/工商页面证据；未观测值保持空值。"
    return {
        "source_platform": "1688",
        "manufacturer_id": f"1688:manufacturer:{member_id}",
        "member_id": member_id,
        "company_id": clean(company.get("company_id")),
        "manufacturer_name": clean(company.get("company_name")) or clean(fallback_name),
        "related_product_ids": [item["product_id"] for item in related_products],
        "product_relation_type": relation_type,
        "shop_url": clean(company.get("shop_url")) or clean(related_products[0].get("product_url")),
        "factory_archive_url": factory_url,
        "subject_qualification_url": factory_url,
        "business_info_url": credit_url,
        "unified_social_credit_code": clean(company.get("unified_social_credit_code")),
        "legal_representative": clean(company.get("legal_representative")),
        "registered_capital": clean(company.get("registered_capital_text")),
        "established_date": established_date,
        "company_type": clean(company.get("company_type")),
        "registration_authority": clean(company.get("registration_authority")),
        "business_term": clean(company.get("business_term")),
        "registered_address": clean(company.get("registered_address")),
        "business_scope": clean(company.get("business_scope")),
        "contact_person": clean(contacts.get("contact_person")),
        "telephone": clean(contacts.get("telephone")),
        "mobile": clean(contacts.get("mobile")),
        "contact_address": clean(contacts.get("address")),
        "main_category": clean(company.get("main_category") or profile.get("business_line")),
        "production_service": clean(factory.get("production_service") or profile.get("production_service")),
        "company_summary": clean(profile.get("company_summary")),
        "brands": factory.get("brands") or [],
        "factory_address": clean(factory.get("factory_address")),
        "factory_area_sqm": area,
        "factory_area_authenticated": factory.get("factory_area_is_authenticated", ""),
        "employee_total_range": clean(factory.get("employee_total_range")),
        "production_line_count": factory.get("production_line_count", ""),
        "production_equipment_count": factory.get("production_equipment_count", credit.get("production_equipment_count", "")),
        "annual_transaction_amount": clean(factory.get("annual_transaction_amount_text") or credit.get("annual_transaction_amount_text")),
        "monthly_output_value": clean(factory.get("monthly_output_value")),
        "processing_methods": factory.get("processing_methods") or [],
        "custom_minimum_order": clean(factory.get("custom_minimum_order")),
        "vat_invoice_available": clean(factory.get("vat_invoice_available")),
        "qualification_tags": tags,
        "certificate_count": certificate_details.get("reported_total", ""),
        "certificates": certificate_items,
        "factory_auth_provider": clean(factory.get("factory_auth_provider")),
        "factory_auth_report_number": clean(factory.get("factory_auth_report_number")),
        "patent_count": patents.get("reported_total", "") if patent_items else "",
        "patents": patent_items,
        "factory_medal": clean(factory.get("factory_medal")),
        "returning_customer_rate": clean(factory.get("returning_customer_rate") or credit.get("returning_customer_rate")),
        "service_response_rate": clean(factory.get("service_response_rate")),
        "on_time_fulfillment_rate": clean(factory.get("on_time_fulfillment_rate")),
        "factory_vr_url": clean(factory.get("factory_vr_url") or profile.get("factory_vr_url")),
        "factory_images": images,
        "factory_videos": videos,
        "observed_at": clean(asset.get("collected_at")),
    }


def excel_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def write_sheet(workbook: Workbook, title: str, fields: dict[str, str], records: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    keys = list(fields)
    sheet.append([fields[key] for key in keys])
    for record in records:
        sheet.append([excel_value(record.get(key, "")) for key in keys])
    sheet.freeze_panes = None
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="E97132")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        if row[0].row == 1:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
    sheet.row_dimensions[1].height = 34
    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 72
    for index, key in enumerate(keys, start=1):
        values = [clean(fields[key])] + [clean(excel_value(item.get(key, ""))) for item in records[:30]]
        width = max(12, min(42, max((len(value) for value in values), default=12) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def validate_workbook(path: Path, product_count: int, manufacturer_count: int, sku_count: int = 0) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if workbook.sheetnames != ["商品信息", "厂家信息", "SKU明细"]:
        raise RuntimeError(f"unexpected sheets: {workbook.sheetnames}")
    expected_rows = {
        "商品信息": product_count + 1,
        "厂家信息": manufacturer_count + 1,
        "SKU明细": sku_count + 1,
    }
    for sheet in workbook.worksheets:
        if sheet.freeze_panes is not None:
            raise RuntimeError(f"{sheet.title} still has freeze panes")
        if sheet.max_row != expected_rows[sheet.title]:
            raise RuntimeError(f"{sheet.title} row count mismatch")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.alignment.horizontal != "center" or cell.alignment.vertical != "center":
                    raise RuntimeError(f"{sheet.title}!{cell.coordinate} is not centered")


def main() -> int:
    parser = argparse.ArgumentParser(description="生成 1688 商品/厂家两表直接交付")
    parser.add_argument("--run-dir", action="append", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--limit", type=int, default=50)
    parser.add_argument("--rich-company-asset", action="append", default=[])
    parser.add_argument(
        "--allow-missing-manufacturer",
        action="store_true",
        help="厂家未采集的商品也导出（厂家字段标待补），实现商品先行交付",
    )
    parser.add_argument("--delivery-id", default="1688_direct_20260716")
    parser.add_argument("--output-prefix", default="1688分类抽样最完整直接版_20260716")
    args = parser.parse_args()

    run_dirs = [Path(item) for item in args.run_dir]
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    assets: dict[str, dict] = {}
    for run_dir in run_dirs:
        for path in sorted((run_dir / "l1" / "company_items").glob("*/company_asset.json")):
            asset = read_json(path)
            member_id = clean(nested(asset, "company", "member_id"))
            if member_id and (
                member_id not in assets or asset_rank(asset) >= asset_rank(assets[member_id])
            ):
                assets[member_id] = asset
    for raw_path in args.rich_company_asset:
        path = Path(raw_path)
        if path.exists():
            asset = read_json(path)
            member_id = clean(nested(asset, "company", "member_id"))
            if member_id and (
                member_id not in assets or asset_rank(asset) >= asset_rank(assets[member_id])
            ):
                assets[member_id] = asset

    selected: list[dict] = []
    seen_offers: set[str] = set()
    product_paths = [
        path
        for run_dir in run_dirs
        for path in sorted((run_dir / "l1" / "product_items").glob("*/product.json"))
    ]
    products_by_offer: dict[str, dict] = {}
    for path in product_paths:
        product = read_json(path)
        offer_id = clean(product.get("offer_id"))
        skus_path = path.parent / "skus.json"
        if skus_path.exists():
            loaded_skus = read_json(skus_path)
            if isinstance(loaded_skus, list):
                product["skus"] = loaded_skus
            else:
                product["skus"] = []
        if offer_id and (
            offer_id not in products_by_offer
            or product_rank(product) >= product_rank(products_by_offer[offer_id])
        ):
            products_by_offer[offer_id] = product
    products = list(products_by_offer.values())
    products.sort(key=lambda item: (clean(item.get("validation_category")), clean(item.get("offer_id"))))
    for product in products:
        offer_id = clean(product.get("offer_id"))
        member_id = clean(product.get("member_id"))
        if not offer_id or offer_id in seen_offers or (
            member_id not in assets and not args.allow_missing_manufacturer
        ):
            continue
        seen_offers.add(offer_id)
        selected.append(product)
        if len(selected) >= args.limit:
            break

    product_records = [
        product_record(product, assets.get(clean(product.get("member_id"))) or {})
        for product in selected
    ]
    sku_records: list[dict] = []
    seen_sku_rows: set[tuple[str, str, str, str]] = set()
    for product in selected:
        offer_id = clean(product.get("offer_id"))
        for item in product.get("skus") or []:
            if not isinstance(item, dict):
                continue
            sku_key = (
                offer_id,
                clean(item.get("sku_name")),
                clean(item.get("sku_price")),
                clean(item.get("stock_quantity")),
            )
            if sku_key in seen_sku_rows:
                continue
            seen_sku_rows.add(sku_key)
            sku_records.append(
                {
                    "product_id": offer_id,
                    "sku_name": clean(item.get("sku_name")),
                    "sku_price": fmt_price(item.get("sku_price")),
                    "sku_price_text": clean(item.get("sku_price_text") or item.get("price_text")),
                    "stock_quantity": clean(item.get("stock_quantity")),
                    "sku_image_url": clean(item.get("sku_image_url") or item.get("image_url")),
                    "collected_at": clean(item.get("collected_at")),
                }
            )
    related: dict[str, list[dict]] = defaultdict(list)
    fallback_names: dict[str, str] = {}
    for record in product_records:
        member_id = clean(record["manufacturer_member_id"])
        related[member_id].append(record)
        fallback_names[member_id] = clean(record["manufacturer_name"])
    manufacturer_records = [
        manufacturer_record(assets[member_id], related[member_id], fallback_names[member_id])
        for member_id in sorted(related)
        if member_id in assets
    ]

    categories = sorted({clean(item.get("youyiquan_category_candidate")) for item in product_records if clean(item.get("youyiquan_category_candidate"))})
    factory_archive_count = sum(bool(item.get("factory_archive_url")) for item in manufacturer_records)
    sparse_factory_count = sum(item.get("data_quality_status") == "factory_archive_sparse_source" for item in manufacturer_records)
    payload = {
        "delivery_id": args.delivery_id,
        "schema_version": "1.1.0",
        "delivery_type": "direct_readable_category_sample",
        "source": "1688",
        "status": "completed_with_source_sparse_fields",
        "description": "商品与厂家通过manufacturer_id关联；厂家按memberId去重；全部厂家均以工厂档案为厂家字段入口，仅输出页面实际公开值，未上传或未披露字段保持空值并注明原因。",
        "sample_summary": {
            "product_count": len(product_records),
            "manufacturer_count": len(manufacturer_records),
            "category_count": len(categories),
            "categories": categories,
            "requested_approximate_product_count": args.limit,
            "factory_archive_visited_count": factory_archive_count,
            "source_sparse_manufacturer_count": sparse_factory_count,
            "live_refresh_status": "completed",
            "missing_manufacturer_product_count": (
                sum(
                    1
                    for record in product_records
                    if clean(record.get("manufacturer_member_id")) not in assets
                )
                if args.allow_missing_manufacturer
                else 0
            ),
        },
        "product_field_labels_zh": PRODUCT_FIELDS,
        "manufacturer_field_labels_zh": MANUFACTURER_FIELDS,
        "sku_field_labels_zh": DELIVERY_SKU_FIELDS,
        "price_summary": {
            "single_price_count": sum(1 for r in product_records if clean(r.get("price_status")) == "single"),
            "range_price_count": sum(1 for r in product_records if clean(r.get("price_status")) == "range"),
            "review_required_count": sum(1 for r in product_records if clean(r.get("price_status")) == "review_required"),
            "missing_price_count": sum(1 for r in product_records if clean(r.get("price_status")) == "missing"),
        },
        "products": product_records,
        "manufacturers": manufacturer_records,
        "skus": sku_records,
    }
    json_path = output_dir / f"{args.output_prefix}.json"
    xlsx_path = output_dir / f"{args.output_prefix}.xlsx"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(workbook, "商品信息", PRODUCT_FIELDS, product_records)
    write_sheet(workbook, "厂家信息", MANUFACTURER_FIELDS, manufacturer_records)
    write_sheet(workbook, "SKU明细", DELIVERY_SKU_FIELDS, sku_records)
    workbook.save(xlsx_path)
    validate_workbook(xlsx_path, len(product_records), len(manufacturer_records), len(sku_records))
    print(json.dumps({"json": str(json_path), "xlsx": str(xlsx_path), "products": len(product_records), "manufacturers": len(manufacturer_records), "categories": len(categories)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
