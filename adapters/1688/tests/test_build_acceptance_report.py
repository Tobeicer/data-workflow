"""验收报告生成器测试（离线）。"""

import json
import sys
from pathlib import Path

SRC_DIR = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(SRC_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared" / "src"))

from build_acceptance_report import (  # noqa: E402
    build_manufacturer_report,
    build_report,
    write_xlsx,
)


def make_asset(member="b2b-m1"):
    return {
        "source_platform": "1688",
        "collected_at": "2026-08-14T10:00:00Z",
        "company": {
            "member_id": member,
            "company_name": "测试厂家有限公司",
            "shop_url": "https://test.1688.com",
            "unified_social_credit_code": "9144TEST",
            "legal_representative": "张三",
            "registered_capital_text": "100万元",
            "established_date": "2020-01-01",
            "company_type": "有限公司",
            "registration_authority": "市监局",
            "business_term": "长期",
            "registered_address": "广东省广州市",
            "business_scope": "玩具制造",
        },
        "contacts": {"contact_person": "李四", "telephone": "13800000000", "mobile": "", "address": "广东省广州市"},
        "company_profile": {"company_summary": "专业玩具", "production_service": "游戏机生产"},
        "company_media": [],
        "factory_snapshots": [],
        "certification_tags": [],
        "certificate_details": {},
        "patent_details": {},
        "field_evidence": [],
        "source_field_observations": [],
        "capture_status": "success",
    }


def test_build_manufacturer_report_46_fields():
    l0 = [{"member_id": "b2b-m1", "pages": [{"page_type": "business_info", "text": "公司名称 测试"}]}]
    report = build_manufacturer_report([make_asset()], l0)
    row = report["manufacturers"][0]
    assert len(row) == 46
    assert row["manufacturer_name"] == "测试厂家有限公司"
    assert row["cross_border_qualification"] == ""
    assert "company_id" not in row
    assert "service_response_rate" not in row
    assert report["summary"][0]["字段填充"] == "22/46"


def test_build_manufacturer_report_carries_quality_issues():
    l0 = [
        {
            "member_id": "b2b-m1",
            "pages": [{"page_type": "business_info", "text": "无标签"}],
            "quality_issues": ["business_info 文本缺'公司名称'标签"],
        }
    ]
    report = build_manufacturer_report([make_asset()], l0)
    assert "公司名称" in report["summary"][0]["合规问题"]


def make_record():
    return {
        "offer_id": "1001",
        "title": "测试娃娃机",
        "member_id": "b2b-1",
        "supplier_name": "测试厂",
        "price_text": "¥34.8-120",
        "price_range_text": "¥34.8-120",
        "sku_dimension": ["颜色"],
        "sku_rows": [
            {"label": "娃娃机【红色】", "priceText": "¥34.8", "stockText": "库存99个", "imageUrl": "https://x/a.jpg"},
            {"label": "娃娃机【蓝色】", "priceText": "¥120", "stockText": "库存1个", "imageUrl": "https://x/b.jpg"},
        ],
        "attributes": {"品牌": "X牌", "材质": "塑料"},
        "image_urls": ["https://x/a.jpg", "https://x/b.jpg"],
        "detail_images": ["https://x/d1.jpg"],
        "video_url": "https://v.example.com/v.mp4",
        "layout_key": "t:found",
        "quality_issues": [],
    }


def test_build_report_views_and_summary():
    report = build_report([make_record()])
    assert len(report["products"]) == 1
    assert report["products"][0]["轮播图数"] == 2
    assert report["products"][0]["详情图数"] == 1
    assert report["summary"][0]["视频"] == "有"
    assert len(report["skus"]) == 2
    red = [s for s in report["skus"] if "红色" in s["sku_name"]][0]
    assert red["sku_price"] == "34.8"
    assert red["spec_parse_status"] == "parsed"
    assert red["stock_quantity"] == "99"


def test_build_report_counts_quality_issues():
    record = make_record()
    record["quality_issues"] = ["sku[0] 价格原文含杂质"]
    report = build_report([record])
    assert report["summary"][0]["合规问题数"] == 1


def test_write_xlsx_three_sheets(tmp_path):
    from openpyxl import load_workbook

    write_xlsx(tmp_path / "a.xlsx", build_report([make_record()]))
    workbook = load_workbook(tmp_path / "a.xlsx")
    assert workbook.sheetnames == ["汇总", "商品", "SKU明细"]
    assert workbook["汇总"].cell(row=2, column=1).value == "1001"
