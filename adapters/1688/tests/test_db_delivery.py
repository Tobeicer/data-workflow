"""db_delivery 映射函数测试（离线）。"""

import sys
from pathlib import Path

SRC_DIR = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(SRC_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared" / "src"))

from db_delivery import (  # noqa: E402
    _manufacturer_row_errors,
    load_registry,
    manufacturer_row,
    product_row,
    registry_keys,
    save_registry,
    sku_rows_from,
    validate_manufacturer_assets,
    validate_products,
    write_snapshot_xlsx,
)


def test_write_snapshot_xlsx_writes_sheets(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    path = write_snapshot_xlsx(
        tmp_path / "s.xlsx",
        [
            ("商品", [{"product_id": "1", "title": "A", "image_urls": ["u1"]}]),
            ("SKU明细", [{"product_id": "1", "sku_name": "s1", "spec_attributes": [{"k": "v"}]}]),
        ],
    )
    assert path.exists()
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["商品", "SKU明细"]
    assert workbook["商品"].cell(row=2, column=1).value == "1"
    # json 列已序列化为文本
    assert workbook["SKU明细"].cell(row=2, column=3).value == '[{"k": "v"}]'


def make_product(**overrides):
    product = {
        "offer_id": "1001",
        "product_url": "https://detail.1688.com/offer/1001.html",
        "title": "测试娃娃机",
        "collected_at": "2026-08-14T10:00:00Z",
        "member_id": "b2b-1",
        "supplier_name": "测试厂",
        "validation_category": "A01",
        "attributes": {"品牌": "X牌", "颜色": "红色"},
        "minimum_order_quantity": "2台",
        "sales_unit": "台",
        "available_stock": "100",
        "sku_count": 2,
        "sku_dimension": "颜色",
        "skus": [
            {
                "sku_name": "娃娃机【红色】",
                "sku_dimension": "颜色",
                "spec_attributes": [{"dimension": "颜色", "value": "红色"}],
                "spec_parse_status": "parsed",
                "sku_price": "100",
                "sku_price_text": "¥100",
                "stock_quantity": "50",
                "sku_image_url": "https://img/a.jpg",
                "collected_at": "2026-08-14T10:00:00Z",
            },
            {
                "sku_name": "娃娃机【蓝色】",
                "sku_dimension": "颜色",
                "spec_attributes": [{"dimension": "颜色", "value": "蓝色"}],
                "spec_parse_status": "parsed",
                "sku_price": "120",
                "sku_price_text": "¥120",
                "stock_quantity": "60",
                "sku_image_url": "https://img/b.jpg",
                "collected_at": "2026-08-14T10:00:00Z",
            },
        ],
    }
    product.update(overrides)
    return product


def test_product_row_price_range_from_skus():
    row = product_row(make_product(), "d1")
    assert row["price_min"] == 100.0
    assert row["price_max"] == 120.0
    assert row["price_status"] == "range"
    assert row["currency"] == "CNY"


def test_product_row_single_price():
    product = make_product()
    product["skus"] = [product["skus"][0]]
    row = product_row(product, "d1")
    assert row["price_status"] == "single"
    assert row["price_min"] == row["price_max"] == 100.0


def test_product_row_missing_price_not_zero():
    product = make_product()
    for s in product["skus"]:
        s["sku_price"] = ""
    row = product_row(product, "d1")
    assert row["price_min"] is None
    assert row["price_status"] == "missing"


def test_product_row_carries_spec_fields():
    row = product_row(make_product(), "d1")
    assert row["sku_dimension"] == "颜色"
    assert row["import_delivery_id"] == "d1"
    assert row["schema_version"] == "1.2.0"
    assert row["manufacturer_id"] == "1688:manufacturer:b2b-1"


def test_product_row_color_sanitized():
    product = make_product()
    product["attributes"] = {"颜色": "红色/蓝色/绿色"}
    row = product_row(product, "d1")
    assert row["color"] == ""


def test_sku_rows_from_carries_spec_fields():
    rows = sku_rows_from(make_product(), "d1")
    assert len(rows) == 2
    assert rows[0]["sku_dimension"] == "颜色"
    assert rows[0]["spec_attributes"] == [{"dimension": "颜色", "value": "红色"}]
    assert rows[0]["spec_parse_status"] == "parsed"
    assert rows[0]["sku_price"] == 100.0
    assert rows[0]["import_delivery_id"] == "d1"


def test_sku_rows_bad_price_is_none():
    product = make_product()
    product["skus"][0]["sku_price"] = "abc"
    rows = sku_rows_from(product, "d1")
    assert rows[0]["sku_price"] is None


def test_manufacturer_row_maps_core_fields():
    asset = {
        "company": {
            "company_name": "广州测试动漫科技有限公司",
            "member_id": "b2b-test-1",
            "province": "广东省",
            "city": "广州市",
            "shop_url": "https://test.1688.com",
            "main_category": "游戏设备",
            "registered_address": "广东省广州市番禺区",
        },
        "contacts": {"contact_person": "张三", "telephone": "", "mobile": "13800000000"},
        "company_profile": {"company_summary": "专注娃娃机", "production_service": "来图加工"},
        "company_media": [
            {"url": "https://img/logo.png", "media_type": "logo"},
            {"url": "https://img/v.mp4", "media_type": "video"},
        ],
        "certification_tags": ["深度认证"],
    }
    row = manufacturer_row(asset, "b1")
    assert row["name"] == "广州测试动漫科技有限公司"
    assert row["member_id"] == "b2b-test-1"
    assert row["region"] == "广东省/广州市/"
    assert row["website"] == "https://test.1688.com"
    assert row["source_url"] == "https://test.1688.com"
    assert row["contact_name"] == "张三"
    assert row["contact_phone"] == "13800000000"
    assert row["status"] == "pending"
    assert row["claim_status"] == "unclaimed"
    assert row["import_batch"] == "b1"
    assert row["logo_url"] == "https://img/logo.png"
    assert row["video_url"] == "https://img/v.mp4"
    assert "深度认证" in row["qualifications"]


def test_manufacturer_row_source_url_falls_back_to_factory_archive():
    asset = _asset(
        name="深圳市鸿裕欣电子科技有限公司",
        source_url="",
        member="b2b-2219031689218c7d6f",
    )
    asset["factory_snapshots"] = [
        {"snapshot_type": "company_header_summary", "source_url": ""},
        {
            "snapshot_type": "factory_archive_page",
            "source_url": "https://sale.1688.com/factory/card.html?memberId=b2b-2219031689218c7d6f",
        },
    ]
    row = manufacturer_row(asset, "b1")
    assert row["source_url"] == (
        "https://sale.1688.com/factory/card.html?memberId=b2b-2219031689218c7d6f"
    )


# -- 去重 -----------------------------------------------------------------


def test_registry_keys_member_priority_and_source_fallback():
    assert registry_keys({"member_id": "b2b-1", "source_url": "https://a.1688.com"}) == [
        "b2b-1",
        "https://a.1688.com",
    ]
    assert registry_keys({"member_id": "", "source_url": "https://a.1688.com"}) == [
        "https://a.1688.com"
    ]


def test_registry_save_and_load_roundtrip(tmp_path):
    path = tmp_path / "reg.json"
    save_registry(path, {"b2b-1": True})
    assert load_registry(path) == {"b2b-1": True}
    assert load_registry(tmp_path / "missing.json") == {}


def test_validate_manufacturers_rejects_duplicate_member_in_batch():
    errors, _ = validate_manufacturer_assets(
        [_asset(member="b2b-1"), _asset(member="b2b-1")], "b1"
    )
    assert any("duplicate member_id" in e for e in errors)


# -- 校验门 -----------------------------------------------------------------


def _asset(name="A厂", source_url="https://a.1688.com", phone="13800000000", contact="张三", member="b2b-a"):
    return {
        "company": {"company_name": name, "member_id": member, "shop_url": source_url},
        "contacts": {"contact_person": contact, "telephone": phone, "mobile": ""},
        "company_profile": {},
        "company_media": [],
        "certification_tags": [],
    }


def test_validate_products_accepts_valid_batch():
    errors, warnings = validate_products([make_product()], "d1")
    assert errors == []


def test_validate_products_rejects_empty_batch():
    errors, _ = validate_products([], "d1")
    assert any("batch empty" in e for e in errors)


def test_validate_products_rejects_empty_title_and_duplicate():
    p1 = make_product()
    p2 = make_product()
    p2["title"] = ""
    errors, _ = validate_products([p1, p2], "d1")
    assert any("title empty" in e for e in errors)
    errors2, _ = validate_products([p1, p1], "d1")
    assert any("duplicate product_id" in e for e in errors2)


def test_validate_products_rejects_bad_price():
    p = make_product()
    p["skus"][0]["sku_price"] = "¥100库存99"  # 拼接残留
    errors, _ = validate_products([p], "d1")
    assert any("not numeric" in e for e in errors)
    p["skus"][0]["sku_price"] = "12000000"
    errors2, _ = validate_products([p], "d1")
    assert any("1千万" in e for e in errors2)


def test_validate_products_rejects_bad_spec_status():
    p = make_product()
    p["skus"][0]["spec_parse_status"] = "weird"
    errors, _ = validate_products([p], "d1")
    assert any("bad spec_parse_status" in e for e in errors)


def test_validate_products_warns_skus_without_dimension():
    p = make_product()
    p["sku_dimension"] = ""
    errors, warnings = validate_products([p], "d1")
    assert errors == []
    assert any("sku_dimension empty" in w for w in warnings)


def test_validate_manufacturers_accepts_valid_batch():
    errors, _ = validate_manufacturer_assets([_asset()], "b1")
    assert errors == []


def test_validate_manufacturers_rejects_invalid_rows():
    errors, _ = validate_manufacturer_assets([_asset(name="")], "b1")
    assert any("without name" in e for e in errors)
    errors2, _ = validate_manufacturer_assets([_asset(source_url="")], "b1")
    assert any("source_url invalid" in e for e in errors2)
    errors3, _ = validate_manufacturer_assets(
        [_asset(), _asset()], "b1"
    )
    assert any("duplicate source_url" in e for e in errors3)
    errors4, _ = validate_manufacturer_assets([_asset(phone="138abc")], "b1")
    assert any("contact_phone invalid" in e for e in errors4)


def test_validate_manufacturers_warns_sparse_contact():
    _, warnings = validate_manufacturer_assets([_asset(contact="")], "b1")
    assert any("contact_name empty" in w for w in warnings)
