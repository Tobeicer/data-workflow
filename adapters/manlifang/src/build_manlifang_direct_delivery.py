# -*- coding: utf-8 -*-
"""漫立方 1688 同格式宽表交付生成器。

读取正式批次 cleaned XLSX + structured JSONL，输出与 1688 交付同结构的
统一 JSON（商品 50 字段 + 厂家 54 字段，扁平宽记录）和同字段中文 Excel。
漫立方为单厂家来源：所有商品 manufacturer_id 指向漫立方一家。

用法：
    python adapters/manlifang/src/build_manlifang_direct_delivery.py
"""

from __future__ import annotations

import argparse
import datetime
import glob
import json
import os
import sys
from collections import Counter, OrderedDict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


SCHEMA_VERSION = "2.3.0"
SOURCE = "manlifang"
MANUFACTURER_ID = "manlifang:manufacturer:manlifang"
MANUFACTURER_NAME = "漫立方"
OBSERVED_AT = "2026-07-10 11:08:14"  # 正式批次 manlifang_full_20260710_110814

# 与 1688 交付 product_field_labels_zh 完全一致的 50 个键（顺序即交付列顺序）
PRODUCT_KEYS = [
    "source_platform", "product_id", "product_url", "title", "observed_at",
    "manufacturer_id", "manufacturer_member_id", "manufacturer_name",
    "manufacturer_relation_type", "manufacturer_relation_status",
    "youyiquan_category_candidate", "product_category", "brand", "model",
    "item_number", "function", "material", "origin", "applicable_people",
    "applicable_age", "applicable_scenarios", "technical_type", "color",
    "ip_authorized", "ccc_configuration_category", "ccc_certificate_code",
    "packaging", "patent_type", "display_price", "minimum_order_quantity",
    "sales_unit", "available_stock", "delivery_commitment",
    "quality_report_number", "pack_length_cm", "pack_width_cm",
    "pack_height_cm", "pack_volume_cm3", "pack_weight_g", "pack_specs_json",
    "main_image_url", "image_urls", "video_url", "detail_content_url",
    "detail_images_json", "service_guarantees", "raw_sku_count",
    "related_product_count", "related_products_json", "other_attributes",
]

# 与 1688 交付 manufacturer_field_labels_zh 完全一致的 54 个键
MANUFACTURER_KEYS = [
    "source_platform", "manufacturer_id", "member_id", "company_id",
    "manufacturer_name", "related_product_ids", "product_relation_type",
    "shop_url", "factory_archive_url", "subject_qualification_url",
    "business_info_url", "unified_social_credit_code", "legal_representative",
    "registered_capital", "established_date", "company_type",
    "registration_authority", "business_term", "registered_address",
    "business_scope", "contact_person", "telephone", "mobile",
    "contact_address", "main_category", "production_service",
    "company_summary", "brands", "factory_address", "factory_area_sqm",
    "factory_area_authenticated", "employee_total_range",
    "production_line_count", "production_equipment_count",
    "annual_transaction_amount", "monthly_output_value", "processing_methods",
    "custom_minimum_order", "vat_invoice_available", "qualification_tags",
    "certificate_count", "certificates", "factory_auth_provider",
    "factory_auth_report_number", "patent_count", "patents", "factory_medal",
    "returning_customer_rate", "service_response_rate",
    "on_time_fulfillment_rate", "factory_vr_url", "factory_images",
    "factory_videos", "observed_at",
]

PRODUCT_LABELS_ZH = {
    "source_platform": "来源平台", "product_id": "商品ID", "product_url": "商品链接",
    "title": "商品标题", "observed_at": "采集时间", "manufacturer_id": "厂家ID",
    "manufacturer_member_id": "厂家memberId", "manufacturer_name": "厂家名称",
    "manufacturer_relation_type": "商品厂家关系", "manufacturer_relation_status": "关联状态",
    "youyiquan_category_candidate": "游艺圈分类候选", "product_category": "产品类别",
    "brand": "品牌", "model": "型号", "item_number": "货号", "function": "功能",
    "material": "材质", "origin": "产地", "applicable_people": "适用人数",
    "applicable_age": "适用年龄", "applicable_scenarios": "适用场景",
    "technical_type": "技术类型", "color": "颜色", "ip_authorized": "是否IP授权",
    "ccc_configuration_category": "3C配置类别", "ccc_certificate_code": "商品3C认证码",
    "packaging": "包装", "patent_type": "专利类型", "display_price": "展示价格",
    "minimum_order_quantity": "起订量", "sales_unit": "销售单位",
    "available_stock": "可售库存", "delivery_commitment": "发货承诺",
    "quality_report_number": "质检报告编号", "pack_length_cm": "包装长(cm)",
    "pack_width_cm": "包装宽(cm)", "pack_height_cm": "包装高(cm)",
    "pack_volume_cm3": "包装体积(cm³)", "pack_weight_g": "包装重量(g)",
    "pack_specs_json": "包装明细(JSON)", "main_image_url": "商品主图",
    "image_urls": "商品图片", "video_url": "商品视频",
    "detail_content_url": "详情内容链接", "detail_images_json": "详情图片(JSON)",
    "service_guarantees": "服务保障", "raw_sku_count": "原始SKU数量",
    "related_product_count": "相关商品数", "related_products_json": "相关商品(JSON)",
    "other_attributes": "其他商品属性",
}

MANUFACTURER_LABELS_ZH = {
    "source_platform": "来源平台", "manufacturer_id": "厂家ID", "member_id": "1688 memberId",
    "company_id": "1688公司ID", "manufacturer_name": "厂家名称",
    "related_product_ids": "关联商品ID", "product_relation_type": "商品厂家关系",
    "shop_url": "店铺链接", "factory_archive_url": "工厂档案链接",
    "subject_qualification_url": "主体资质入口", "business_info_url": "工商详情链接",
    "unified_social_credit_code": "统一社会信用代码", "legal_representative": "法定代表人",
    "registered_capital": "注册资本", "established_date": "成立日期",
    "company_type": "企业类型", "registration_authority": "登记机关",
    "business_term": "营业期限", "registered_address": "注册地址",
    "business_scope": "经营范围", "contact_person": "联系人", "telephone": "联系电话",
    "mobile": "手机", "contact_address": "联系地址", "main_category": "主营类目",
    "production_service": "主营产品/服务", "company_summary": "公司简介",
    "brands": "自有品牌", "factory_address": "工厂地址",
    "factory_area_sqm": "工厂面积（㎡）", "factory_area_authenticated": "工厂面积已认证",
    "employee_total_range": "员工规模", "production_line_count": "生产线数量",
    "production_equipment_count": "生产设备数量", "annual_transaction_amount": "年交易额",
    "monthly_output_value": "月产值", "processing_methods": "加工方式",
    "custom_minimum_order": "定制起订量", "vat_invoice_available": "可开增值税发票",
    "qualification_tags": "工厂能力/认证标签", "certificate_count": "资质证书数量",
    "certificates": "资质证书明细", "factory_auth_provider": "深度认证机构",
    "factory_auth_report_number": "认证报告编号", "patent_count": "专利数量",
    "patents": "专利明细", "factory_medal": "工厂牌级",
    "returning_customer_rate": "回头率", "service_response_rate": "服务响应率",
    "on_time_fulfillment_rate": "准时履约率", "factory_vr_url": "工厂VR展厅",
    "factory_images": "工厂图片", "factory_videos": "工厂视频", "observed_at": "采集时间",
}


def read_cleaned_sheet(xlsx_path, sheet_name):
    """读取清洗 XLSX 的指定 sheet，返回 (表头列表, 行字典列表)。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    records = []
    for row in rows:
        if row is None or all(v is None or v == "" for v in row):
            continue
        records.append(dict(zip(header, row)))
    wb.close()
    return header, records


def load_jsonl(path):
    with open(path, encoding="utf-8") as f:
        return [json.loads(line) for line in f]


SPU_CORE_FIELDS = {
    "NAME", "PRICE", "MEMBER_PRICE", "IMAGE", "RETAIL_PRICE",
    "AVAIL_QTY", "CODE", "SPEC_NAME", "SPEC_AVAIL_QTY",
}

# structured JSONL 漏提取、但 raw 响应中真实存在的字段
RAW_STATIC_KEYS = ("aliasName", "specNo", "auxUomName", "auxUomConversionRate")


def load_raw_static_extra(raw_static_dir):
    """从 raw static_detail 响应补读 structured JSONL 未提取的真实字段。"""
    extra = {}
    for fp in glob.glob(os.path.join(raw_static_dir, "*.json")):
        try:
            with open(fp, encoding="utf-8") as f:
                d = json.load(f)
        except Exception:
            continue
        pid = d.get("productId")
        if pid is None:
            continue
        out = {}
        for k in RAW_STATIC_KEYS:
            v = d.get(k)
            if v not in (None, "", []):
                out[k] = v
        if out:
            extra[str(pid)] = out
    return extra


def build_products(cleaned_records, image_by_code, spu_by_product, static_by_product,
                   v2_by_code, raw_extra_by_product=None, listing_by_product=None):
    """组装 50 字段商品宽记录。

    cleaned_records: 商品清洗主表记录
    image_by_code : {product_code: {"main": [url...], "detail": [url...]}}
    spu_by_product: {product_id: {字段名: 值}}
    """
    products = []
    stats = Counter()
    for rec in cleaned_records:
        code = str(rec.get("source_product_code") or "").strip()
        pid = rec.get("source_product_id")
        imgs = image_by_code.get(code, {})
        prim_urls = imgs.get("main", [])
        detail_urls = imgs.get("detail", [])
        spu = spu_by_product.get(str(pid), {})
        st = static_by_product.get(str(pid), {})
        raw_extra = (raw_extra_by_product or {}).get(str(pid), {})
        listing_extra = (listing_by_product or {}).get(str(pid), {})

        pmin, pmax = rec.get("price_min"), rec.get("price_max")
        if pmin is not None and pmax is not None and float(pmin) == float(pmax):
            display_price = str(pmin)
        elif pmin is not None and pmax is not None:
            display_price = f"{pmin}~{pmax}"
        else:
            display_price = ""

        other = {}
        for k, v in spu.items():
            if k not in SPU_CORE_FIELDS and v not in (None, ""):
                other[k] = v
        other["月销量"] = other.pop("MONTH_SALE_AMOUNT", None)
        if other.get("月销量") is None:
            other.pop("月销量", None)
        if listing_extra.get("displayName"):
            other["展示名"] = listing_extra["displayName"]
        logistics = st.get("logisticsModeList") or []
        if logistics:
            other["物流方式"] = "、".join(logistics)
        if st.get("baseUomName"):
            other["基础单位"] = st["baseUomName"]
        if st.get("mshopProductStatusEnum"):
            other["商品状态"] = st["mshopProductStatusEnum"]
        if st.get("isMultiSpecEnabled") is not None:
            other["是否多规格"] = "是" if st["isMultiSpecEnabled"] else "否"
        if st.get("isProductSoldOut") is not None:
            other["是否售罄"] = "是" if st["isProductSoldOut"] else "否"
        if raw_extra.get("aliasName"):
            other["别名"] = raw_extra["aliasName"]
        if raw_extra.get("specNo"):
            other["规格编号"] = raw_extra["specNo"]
        if raw_extra.get("auxUomName"):
            other["辅助单位"] = raw_extra["auxUomName"]
            rate = raw_extra.get("auxUomConversionRate")
            if rate not in (None, ""):
                other["辅助单位换算率"] = rate
        stamp = listing_extra.get("onShelfStamp")
        if stamp:
            try:
                other["上架时间"] = datetime.datetime.fromtimestamp(
                    float(stamp) / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                other["上架时间"] = str(stamp)
        if rec.get("description_candidate"):
            other["来源描述"] = str(rec["description_candidate"])[:500]
        if not other:
            other = {}

        products.append(OrderedDict([
            ("source_platform", "漫立方"),
            ("product_id", pid if pid is not None else ""),
            ("product_url", rec.get("source_link") or ""),
            ("title", rec.get("normalized_name") or rec.get("original_name") or ""),
            ("observed_at", OBSERVED_AT),
            ("manufacturer_id", MANUFACTURER_ID),
            ("manufacturer_member_id", "manlifang"),
            ("manufacturer_name", MANUFACTURER_NAME),
            ("manufacturer_relation_type", "factory_supplier_of"),
            ("manufacturer_relation_status", "source_factory_linked"),
            ("youyiquan_category_candidate", v2_by_code.get(code, "")),
            ("product_category", rec.get("real_category") or ""),
            ("brand", "漫立方"),
            ("model", rec.get("model_candidate") or ""),
            ("item_number", code),
            ("function", ""), ("material", ""), ("origin", ""),
            ("applicable_people", ""), ("applicable_age", ""),
            ("applicable_scenarios", ""), ("technical_type", ""), ("color", ""),
            ("ip_authorized", ""), ("ccc_configuration_category", ""),
            ("ccc_certificate_code", ""), ("packaging", ""), ("patent_type", ""),
            ("display_price", display_price),
            ("minimum_order_quantity", ""),
            ("sales_unit", rec.get("sales_unit") or ""),
            ("available_stock", rec.get("stock_qty_snapshot") if rec.get("stock_qty_snapshot") is not None else ""),
            ("delivery_commitment", ""), ("quality_report_number", ""),
            ("pack_length_cm", ""), ("pack_width_cm", ""), ("pack_height_cm", ""),
            ("pack_volume_cm3", ""), ("pack_weight_g", ""), ("pack_specs_json", ""),
            ("main_image_url", prim_urls[0] if prim_urls else ""),
            ("image_urls", prim_urls),
            ("video_url", ""), ("detail_content_url", ""),
            ("detail_images_json", json.dumps(detail_urls, ensure_ascii=False) if detail_urls else ""),
            ("service_guarantees", ""),
            ("raw_sku_count", 0),
            ("related_product_count", 0),
            ("related_products_json", ""),
            ("other_attributes", other),
        ]))
        if prim_urls:
            stats["has_main"] += 1
        if detail_urls:
            stats["has_detail"] += 1
        stats["main_urls"] += len(prim_urls)
        stats["detail_urls"] += len(detail_urls)
    return products, stats


def build_manufacturer(all_product_ids, main_categories):
    """组装单厂家（漫立方）54 字段宽记录。"""
    return OrderedDict([
        ("source_platform", "漫立方"),
        ("manufacturer_id", MANUFACTURER_ID),
        ("member_id", ""), ("company_id", ""),
        ("manufacturer_name", MANUFACTURER_NAME),
        ("related_product_ids", all_product_ids),
        ("product_relation_type", "factory_supplier_of"),
        ("shop_url", "漫立方配件商城（微信小程序）"),
        ("factory_archive_url", ""), ("subject_qualification_url", ""),
        ("business_info_url", ""), ("unified_social_credit_code", ""),
        ("legal_representative", ""), ("registered_capital", ""),
        ("established_date", ""), ("company_type", ""),
        ("registration_authority", ""), ("business_term", ""),
        ("registered_address", ""), ("business_scope", ""),
        ("contact_person", ""), ("telephone", ""), ("mobile", ""),
        ("contact_address", ""),
        ("main_category", " | ".join(sorted(main_categories)) if main_categories else ""),
        ("production_service", "游艺设备、配件、耗材供应"),
        ("company_summary", "漫立方配件城：行业较大的配件商城，通过邀约入驻提供配件数据。"),
        ("brands", ["漫立方"]),
        ("factory_address", ""), ("factory_area_sqm", ""),
        ("factory_area_authenticated", ""), ("employee_total_range", ""),
        ("production_line_count", ""), ("production_equipment_count", ""),
        ("annual_transaction_amount", ""), ("monthly_output_value", ""),
        ("processing_methods", ""), ("custom_minimum_order", ""),
        ("vat_invoice_available", ""), ("qualification_tags", ""),
        ("certificate_count", ""), ("certificates", ""),
        ("factory_auth_provider", ""), ("factory_auth_report_number", ""),
        ("patent_count", ""), ("patents", ""), ("factory_medal", ""),
        ("returning_customer_rate", ""), ("service_response_rate", ""),
        ("on_time_fulfillment_rate", ""), ("factory_vr_url", ""),
        ("factory_images", ""), ("factory_videos", ""),
        ("observed_at", OBSERVED_AT),
    ])


def to_excel_cell(value):
    """列表/字典字段在 Excel 中以 JSON 字符串表示（与 1688 交付一致）。"""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_xlsx(path, products, manufacturers):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "商品信息"
    ws1.append([PRODUCT_LABELS_ZH[k] for k in PRODUCT_KEYS])
    for p in products:
        ws1.append([to_excel_cell(p[k]) for k in PRODUCT_KEYS])

    ws2 = wb.create_sheet("厂家信息")
    ws2.append([MANUFACTURER_LABELS_ZH[k] for k in MANUFACTURER_KEYS])
    for m in manufacturers:
        ws2.append([to_excel_cell(m[k]) for k in MANUFACTURER_KEYS])

    # 冻结表头，便于查看
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="漫立方 1688 同格式宽表交付生成")
    parser.add_argument("--run-dir", default=r"runtime/runs/manlifang/manlifang_full_20260710_110814",
                        help="正式批次目录（相对仓库根或绝对路径）")
    parser.add_argument("--out-dir", default="", help="输出目录，默认 deliveries/manlifang/manlifang_direct_<日期>")
    args = parser.parse_args()

    run_dir = args.run_dir
    if not os.path.isabs(run_dir):
        run_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), run_dir)
    structured = os.path.join(run_dir, "structured")
    xlsx_glob = glob.glob(os.path.join(run_dir, "cleaned", "*.xlsx"))
    if not xlsx_glob:
        sys.exit(f"cleaned XLSX 不存在: {run_dir}\\cleaned")
    cleaned_xlsx = xlsx_glob[0]

    print(f"[1/4] 读取清洗主数据: {cleaned_xlsx}")
    _, cleaned_records = read_cleaned_sheet(cleaned_xlsx, "商品清洗主表")
    _, image_records = read_cleaned_sheet(cleaned_xlsx, "图片映射")
    _, relation_records = read_cleaned_sheet(cleaned_xlsx, "商品类目关系")
    print(f"      商品清洗主表 {len(cleaned_records)} 条，图片映射 {len(image_records)} 条，商品类目关系 {len(relation_records)} 条")

    image_by_code = {}
    for im in image_records:
        code = str(im.get("product_code") or "").strip()
        url = im.get("original_url")
        if not code or not url:
            continue
        role = str(im.get("image_role") or "main")
        image_by_code.setdefault(code, {}).setdefault(role, []).append(url)
    print(f"      有图片映射的商品数: {len(image_by_code)}")

    v2_by_code = {}
    for rr in relation_records:
        code = str(rr.get("source_product_code") or "").strip()
        cand = str(rr.get("v2_category_candidate") or "").strip()
        if code and cand and code not in v2_by_code:
            v2_by_code[code] = cand
    print(f"      有游艺圈分类候选的商品数: {len(v2_by_code)}")

    print("[2/4] 读取结构化 JSONL")
    spu_by_product = {}
    for rec in load_jsonl(os.path.join(structured, "spu_details.jsonl")):
        fields = {}
        for fld in rec.get("fields", []):
            fields[fld.get("name")] = fld.get("value")
        spu_by_product[str(rec.get("productId"))] = fields
    static_by_product = {}
    for rec in load_jsonl(os.path.join(structured, "static_details.jsonl")):
        static_by_product[str(rec.get("productId"))] = rec
    listing_by_product = {}
    for rec in load_jsonl(os.path.join(structured, "products.jsonl")):
        extra = {}
        rp, rl = rec.get("raw_product"), rec.get("raw_listing")
        for o in (rl, rp):
            if isinstance(o, dict) and o.get("displayName"):
                extra["displayName"] = o["displayName"]
                break
        for o in (rl, rp):
            if isinstance(o, dict) and o.get("onShelfStamp"):
                extra["onShelfStamp"] = o["onShelfStamp"]
                break
        if extra:
            listing_by_product[str(rec.get("product_id"))] = extra
    print(f"      从 products.jsonl 补读上架时间: {sum(1 for v in listing_by_product.values() if 'onShelfStamp' in v)} 个，展示名: {sum(1 for v in listing_by_product.values() if 'displayName' in v)} 个")
    raw_static_dir = os.path.join(run_dir, "raw", "responses", "static_detail")
    raw_extra_by_product = load_raw_static_extra(raw_static_dir)
    print(f"      从 raw 补读别名/规格编号/辅助单位字段: {len(raw_extra_by_product)} 个商品")
    main_categories = set()
    for cat in load_jsonl(os.path.join(structured, "categories.jsonl")):
        if cat.get("depth") == 0 and cat.get("name"):
            main_categories.add(cat["name"])

    print("[3/4] 组装商品与厂家宽记录")
    products, stats = build_products(cleaned_records, image_by_code, spu_by_product,
                                     static_by_product, v2_by_code, raw_extra_by_product,
                                     listing_by_product)
    all_pids = [str(p["product_id"]) for p in products if p["product_id"] != ""]
    manufacturers = [build_manufacturer(all_pids, main_categories)]

    out_dir = args.out_dir or os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
        "deliveries", "manlifang", "manlifang_direct_20260810")
    os.makedirs(out_dir, exist_ok=True)
    delivery_id = "manlifang_20260810"
    json_path = os.path.join(out_dir, f"漫立方全量_{delivery_id.split('_')[1]}.json")
    xlsx_path = os.path.join(out_dir, f"漫立方全量_{delivery_id.split('_')[1]}.xlsx")

    payload = {
        "delivery_id": delivery_id,
        "schema_version": SCHEMA_VERSION,
        "delivery_type": "standard",
        "source": SOURCE,
        "status": "review_only",
        "description": "正式交付:商品 50 列 + 厂家 54 列(漫立方单厂家);全量原始数据保留 L1-L2",
        "sample_summary": {
            "product_count": len(products),
            "manufacturer_count": len(manufacturers),
            "product_field_count": len(PRODUCT_KEYS),
            "manufacturer_field_count": len(MANUFACTURER_KEYS),
        },
        "product_field_labels_zh": dict(PRODUCT_LABELS_ZH),
        "manufacturer_field_labels_zh": dict(MANUFACTURER_LABELS_ZH),
        "products": products,
        "manufacturers": manufacturers,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    print(f"[4/4] 生成 Excel: {xlsx_path}")
    write_xlsx(xlsx_path, products, manufacturers)

    print("=" * 60)
    print(f"交付 ID : {delivery_id}")
    print(f"商品数  : {len(products)}（有主图 {stats['has_main']}，有详情图 {stats['has_detail']}）")
    print(f"主图 URL: {stats['main_urls']} 张 | 详情图 URL: {stats['detail_urls']} 张")
    print(f"厂家数  : {len(manufacturers)}（{MANUFACTURER_NAME}，关联商品 {len(all_pids)}）")
    print(f"主营类目: {' | '.join(sorted(main_categories))}")
    print(f"JSON 输出: {json_path}")
    print(f"Excel 输出: {xlsx_path}")


if __name__ == "__main__":
    main()
