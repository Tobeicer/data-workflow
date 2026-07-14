from __future__ import annotations

import argparse
import json
import os
import re
import shutil
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


SOURCE_PROVIDER_NAME = "漫立方"
SOURCE_LINK_LABEL = "漫立方配件商城（微信小程序）"


PLATFORM_CATEGORY_RULES = [
    (
        "币器/投币器",
        r"投币器|退币器|币器|投珠器|退珠器",
        r"投币器类|票币器材.*(?:投币|退币)|退币器",
    ),
    (
        "主板/程序板",
        r"主板|程序板|电路板|控制板|控制器|I/?O板|PCB板|驱动板",
        r"主板|程序板|电路板|控制板|游戏机套件.*主板",
    ),
    (
        "摇杆按钮",
        r"摇杆|按钮|按键|微动|方向盘|脚踏开关|操作杆",
        r"摇杆|按钮|按键|控台配件.*微动|方向盘",
    ),
    (
        "读卡器",
        r"读卡器|刷卡器|读卡头|感应卡|IC卡|ID卡",
        r"读卡器|刷卡器|读卡头",
    ),
    (
        "屏幕显示",
        r"显示器|液晶|屏幕|触摸屏|数码管|LED显示|显示屏",
        r"显示器|屏幕|液晶|数码管|显示屏",
    ),
    (
        "支付模块",
        r"支付|扫码|二维码|微信|支付宝|聚合支付",
        r"支付|扫码|二维码",
    ),
    (
        "彩票纸/耗材",
        r"彩票纸|票纸|热敏纸|打印纸|色带|彩票盒|出票器|出卡器|卡片",
        r"彩票纸|票据|打印耗材|出票|卡片机",
    ),
    (
        "礼品机配件",
        r"娃娃机爪|机爪|天车|剪刀机|礼品机|爪套|取物门|礼品出口|抓烟机",
        r"娃娃机爪|天车配件|礼品机配件|娃娃机.*(?:机爪|爪套)",
    ),
    (
        "电源线材",
        r"电源|变压器|滤波器|插座|排插|电源线|信号线|连接线|数据线|转接线|线束|接插件|端子|船型开关",
        r"电源盒|电源线|信号线|排插|插座|变压器|滤波器|灯光线材",
    ),
    (
        "礼品",
        r"公仔|毛绒|礼品|玩具|钥匙扣|挂件|扭蛋球|玻璃珠|弹珠|球饼|(?:篮球|曲棍球)\s*:",
        r"礼品|公仔|毛绒|球类类目|扭蛋球",
    ),
]

NON_CORE_PATH_RE = re.compile(
    r"工具类目|劳保化工|焊机配件|空压机|螺丝|螺母|五金类目|耗材类目|五金机箱配件|五金塑胶",
    re.I,
)
NON_CORE_NAME_RE = re.compile(
    r"空压机|焊机|焊枪|钻头|麻花钻|扳手|螺丝刀|丝攻|套筒|喷漆|胶水|玻璃胶|水鞋|手套|自封袋|打包带|扎带",
    re.I,
)
WHOLE_MACHINE_RE = re.compile(
    r"整机|游戏机台|游艺机|(?:赛车|摩托|篮球|曲棍球|钓鱼|打地鼠|跳舞|射击|娃娃|扭蛋|摇摆|格斗)游戏机",
    re.I,
)
PART_HINT_RE = re.compile(
    r"配件|主板|按钮|按键|摇杆|线|板|灯|轮|爪|锁|螺丝|电源|马达|微动|光眼|投币|铁片|套件",
    re.I,
)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number} is not a JSON object")
            rows.append(value)
    return rows


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_product_code(value: Any) -> str:
    text = clean_text(value)
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    return text.strip("._-") or "UNKNOWN"


def normalize_name(value: Any) -> str:
    text = clean_text(value).replace("：", ":")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*:\s*", ": ", text)
    return text.strip()


def extract_model_candidate(name: str) -> str:
    candidates = re.findall(
        r"(?<![A-Za-z0-9])(?:[A-Za-z]{1,8}[-/]?[A-Za-z0-9]+(?:[-/][A-Za-z0-9]+)*|[A-Z]{1,4}\d{2,})(?![A-Za-z0-9])",
        name,
    )
    unique: list[str] = []
    for value in candidates:
        value = value.strip()
        if value and value not in unique:
            unique.append(value)
    return " ".join(unique[:4])[:100]


def normalize_role(value: Any) -> str:
    role = clean_text(value).lower()
    if role in {"main", "primary", "cover", "image"}:
        return "main"
    if role in {"detail", "description"}:
        return "detail"
    return "other"


def normalized_suffix(value: Any) -> str:
    suffix = Path(clean_text(value)).suffix.lower()
    if suffix in {".jpeg", ".jpe"}:
        return ".jpg"
    if suffix in {".jpg", ".png", ".webp", ".gif", ".bmp", ".avif"}:
        return suffix
    return ".jpg"


def logical_image_name(product_code: Any, role: Any, sequence: int, sha256: Any, suffix: Any) -> str:
    code = normalize_product_code(product_code)
    image_role = normalize_role(role)
    digest = clean_text(sha256).lower()[:8] or "nohash00"
    return f"MLF_{code}_{image_role}_{int(sequence):02d}_{digest}{normalized_suffix(suffix)}"


def assign_logical_image_names(
    rows: Iterable[dict[str, Any]], product_code_by_id: dict[str, str]
) -> list[dict[str, Any]]:
    counters: defaultdict[tuple[str, str], int] = defaultdict(int)
    result: list[dict[str, Any]] = []
    for source_row in rows:
        row = dict(source_row)
        product_id = clean_text(row.get("product_id"))
        product_code = product_code_by_id.get(product_id, f"UNKNOWN_{product_id or 'NO_ID'}")
        role = normalize_role(row.get("image_role"))
        counters[(product_code, role)] += 1
        sequence = counters[(product_code, role)]
        local_file = clean_text(row.get("local_file"))
        name = logical_image_name(product_code, role, sequence, row.get("sha256"), local_file)
        row.update(
            {
                "product_code": product_code,
                "image_role": role,
                "image_sequence": sequence,
                "logical_file_name": name,
                "logical_relative_path": str(Path("cleaned") / "images" / product_code / name).replace("\\", "/"),
            }
        )
        result.append(row)
    return result


def classify_product(name: Any, source_paths: Iterable[Any]) -> dict[str, str]:
    clean_name = normalize_name(name)
    path_text = " | ".join(clean_text(path) for path in source_paths if clean_text(path))

    if WHOLE_MACHINE_RE.search(clean_name) and not PART_HINT_RE.search(clean_name):
        return {
            "business_type": "整机产品候选",
            "platform_category": "",
            "classification_confidence": "中",
            "classification_reason": "名称命中整机/游艺设备规则，需转入产品库复核",
            "review_status": "needs_review",
        }

    if NON_CORE_NAME_RE.search(clean_name):
        return {
            "business_type": "非核心设备/工具候选",
            "platform_category": "",
            "classification_confidence": "中",
            "classification_reason": "名称命中工具、耗材或辅助设备规则",
            "review_status": "needs_review",
        }

    matches: list[tuple[int, int, str, bool, bool]] = []
    for index, (category, name_pattern, path_pattern) in enumerate(PLATFORM_CATEGORY_RULES):
        name_hit = bool(re.search(name_pattern, clean_name, flags=re.I))
        path_hit = bool(re.search(path_pattern, path_text, flags=re.I))
        score = int(name_hit) * 2 + int(path_hit)
        if score:
            matches.append((score, -index, category, name_hit, path_hit))

    if matches:
        _, _, category, name_hit, path_hit = max(matches)
        if category == "礼品":
            business_type = "礼品/玩法道具候选"
        else:
            business_type = "配件候选"
        if name_hit and path_hit:
            confidence = "高"
            reason = f"名称和来源类目均命中{category}规则"
            review_status = "auto_candidate"
        else:
            confidence = "中"
            hit_source = "名称" if name_hit else "来源类目"
            reason = f"仅{hit_source}命中{category}规则，需复核"
            review_status = "needs_review"
        return {
            "business_type": business_type,
            "platform_category": category,
            "classification_confidence": confidence,
            "classification_reason": reason,
            "review_status": review_status,
        }

    core_context = bool(
        re.search(
            r"机台套件|游戏机套件|游戏机台|电子器材|控台配件|马达类目|喇叭功放|投币器类|按钮类目|摇杆/手柄|票币器材|电源盒|信号线类目|电源线类目|装饰灯光",
            path_text,
            flags=re.I,
        )
    )
    if NON_CORE_PATH_RE.search(path_text) and not core_context:
        return {
            "business_type": "非核心设备/工具候选",
            "platform_category": "",
            "classification_confidence": "中",
            "classification_reason": "来源类目命中工具、耗材、五金或辅助设备规则",
            "review_status": "needs_review",
        }

    if core_context:
        return {
            "business_type": "配件候选",
            "platform_category": "",
            "classification_confidence": "低",
            "classification_reason": "来源类目显示为游艺配件，但 v2 分类无安全对应项",
            "review_status": "needs_review",
        }

    return {
        "business_type": "待复核",
        "platform_category": "",
        "classification_confidence": "低",
        "classification_reason": "名称和来源类目不足以确定平台业务类型",
        "review_status": "needs_review",
    }


def index_by(rows: Iterable[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = clean_text(row.get(key))
        if value:
            result[value] = row
    return result


def first_nonempty(*values: Any) -> Any:
    for value in values:
        if value not in (None, "", [], {}):
            return value
    return ""


def numeric_or_blank(value: Any) -> float | int | str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    return int(number) if number.is_integer() else round(number, 4)


def spu_fields(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field in row.get("fields") or []:
        if isinstance(field, dict) and clean_text(field.get("name")):
            result[clean_text(field.get("name"))] = field.get("value", "")
    return result


def joined_paths(paths: Iterable[Any]) -> str:
    unique: list[str] = []
    for path in paths:
        value = clean_text(path).strip("^").replace("^", "/")
        if value and value not in unique:
            unique.append(value)
    return " | ".join(unique)


def choose_primary_source_category(paths: Iterable[Any]) -> str:
    categories = [value for value in joined_paths(paths).split(" | ") if value]
    if not categories:
        return ""
    compatibility_roots = {"新品上新", "机台套件", "游戏机套件", "游戏机台", "精品配件"}

    def score(category: str) -> tuple[int, int, str]:
        parts = category.split("/")
        specialist = int(parts[0] not in compatibility_roots)
        return specialist, len(parts), category

    return max(categories, key=score)


def build_image_rows(batch_dir: Path) -> list[dict[str, Any]]:
    discovered = read_jsonl(batch_dir / "discovered_image_urls.jsonl")
    downloaded = read_jsonl(batch_dir / "downloaded_image_manifest.jsonl")
    download_by_url = {
        clean_text(row.get("url")): row
        for row in downloaded
        if clean_text(row.get("url")) and row.get("status") == "downloaded"
    }
    rows: list[dict[str, Any]] = []
    for relation in discovered:
        url = clean_text(
            relation.get("url") or relation.get("canonical_url") or relation.get("original_url")
        )
        download = download_by_url.get(url, {})
        rows.append(
            {
                "product_id": clean_text(relation.get("product_id")),
                "image_role": normalize_role(relation.get("image_role")),
                "source_endpoint": clean_text(relation.get("source_endpoint")),
                "json_path": clean_text(relation.get("json_path")),
                "original_url": clean_text(relation.get("original_url") or url),
                "canonical_url": clean_text(relation.get("canonical_url") or url),
                "download_status": clean_text(download.get("status")) or "missing",
                "content_type": clean_text(download.get("content_type")),
                "bytes": numeric_or_blank(download.get("bytes")),
                "sha256": clean_text(download.get("sha256")),
                "local_file": clean_text(download.get("local_file")),
                "download_error": clean_text(download.get("error")),
            }
        )
    return rows


def enrich_image_rows(batch_dir: Path, rows: list[dict[str, Any]]) -> None:
    dimensions: dict[str, tuple[int, int, str]] = {}
    for row in rows:
        local_file = clean_text(row.get("local_file"))
        if not local_file:
            width = height = 0
            error = "missing local file"
        elif local_file in dimensions:
            width, height, error = dimensions[local_file]
        else:
            source = batch_dir / local_file
            try:
                with Image.open(source) as image:
                    width, height = image.size
                error = ""
            except (OSError, ValueError) as exc:
                width = height = 0
                error = str(exc)
            dimensions[local_file] = (width, height, error)
        size = numeric_or_blank(row.get("bytes"))
        size_value = float(size) if size != "" else 0.0
        row["width_px"] = width
        row["height_px"] = height
        row["long_side_px"] = max(width, height)
        row["meets_500kb"] = bool(size_value and size_value <= 500 * 1024)
        row["meets_1000px"] = bool(width and height and max(width, height) <= 1000)
        row["image_read_error"] = error
        row["import_image_ready"] = bool(
            row.get("download_status") == "downloaded"
            and row["meets_500kb"]
            and row["meets_1000px"]
            and not error
        )


def reset_generated_image_dir(batch_dir: Path) -> Path:
    image_dir = (batch_dir / "cleaned" / "images").resolve()
    batch_resolved = batch_dir.resolve()
    if batch_resolved not in image_dir.parents or image_dir.name != "images":
        raise RuntimeError(f"unsafe generated image path: {image_dir}")
    if image_dir.exists():
        shutil.rmtree(image_dir)
    image_dir.mkdir(parents=True)
    return image_dir


def create_logical_images(batch_dir: Path, rows: list[dict[str, Any]]) -> Counter:
    image_dir = reset_generated_image_dir(batch_dir)
    modes: Counter = Counter()
    for row in rows:
        local_file = clean_text(row.get("local_file"))
        if not local_file:
            row["logical_link_status"] = "missing_source"
            continue
        source = (batch_dir / local_file).resolve()
        target = image_dir / clean_text(row.get("product_code")) / clean_text(row.get("logical_file_name"))
        target.parent.mkdir(parents=True, exist_ok=True)
        try:
            os.link(source, target)
            mode = "hardlink"
        except OSError:
            shutil.copy2(source, target)
            mode = "copy"
        row["logical_link_status"] = mode
        modes[mode] += 1
    return modes


def build_product_rows(
    products: list[dict[str, Any]],
    static_rows: list[dict[str, Any]],
    dynamic_rows: list[dict[str, Any]],
    spu_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
    batch_id: str,
) -> list[dict[str, Any]]:
    static_by_id = index_by(static_rows, "productId")
    dynamic_by_id = index_by(dynamic_rows, "productId")
    spu_by_id = index_by(spu_rows, "productId")
    images_by_product: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for image in image_rows:
        images_by_product[clean_text(image.get("product_id"))].append(image)

    result: list[dict[str, Any]] = []
    for product in products:
        product_id = clean_text(product.get("product_id"))
        code = normalize_product_code(product.get("product_code"))
        static = static_by_id.get(product_id, {})
        dynamic = dynamic_by_id.get(product_id, {})
        spu = spu_by_id.get(product_id, {})
        fields = spu_fields(spu)
        name_original = clean_text(first_nonempty(static.get("name"), product.get("product_name"), fields.get("NAME")))
        name_normalized = normalize_name(name_original)
        paths = product.get("source_tree_paths") or []
        all_real_categories = joined_paths(paths)
        real_category = choose_primary_source_category(paths)
        classification = classify_product(name_normalized, paths)
        price = numeric_or_blank(
            first_nonempty(
                dynamic.get("retailPrice"),
                dynamic.get("price"),
                static.get("vRetailPriceForRetailUom"),
                (product.get("raw_listing") or {}).get("vRetailPrice"),
                fields.get("PRICE"),
            )
        )
        stock = numeric_or_blank(first_nonempty(dynamic.get("availQty"), fields.get("AVAIL_QTY")))
        product_images = images_by_product.get(product_id, [])
        main_images = [row for row in product_images if row.get("image_role") == "main"]
        detail_images = [row for row in product_images if row.get("image_role") == "detail"]
        noncompliant_count = sum(not bool(row.get("import_image_ready")) for row in product_images)
        description_original = clean_text(static.get("displayDescription"))
        description_candidate = description_original or "；".join(
            part
            for part in (
                f"商品：{name_normalized}" if name_normalized else "",
                f"来源分类：{joined_paths(paths)}" if paths else "",
                f"销售单位：{clean_text(static.get('defaultUomName'))}" if clean_text(static.get("defaultUomName")) else "",
            )
            if part
        )

        quality_flags: list[str] = []
        if not product_images:
            quality_flags.append("缺少图片")
        if not description_original:
            quality_flags.append("缺少来源描述")
        if stock == 0:
            quality_flags.append("库存为0")
        if noncompliant_count:
            quality_flags.append(f"{noncompliant_count}张图片待压缩")
        if classification["review_status"] == "needs_review":
            quality_flags.append("业务分类待复核")

        if classification["review_status"] == "needs_review" or not product_images:
            priority = "P0"
        elif quality_flags:
            priority = "P1"
        else:
            priority = "P2"

        result.append(
            {
                "source_system": "manlifang",
                "capture_batch": batch_id,
                "source_product_id": product_id,
                "source_product_code": code,
                "original_name": name_original,
                "normalized_name": name_normalized,
                "model_candidate": extract_model_candidate(name_normalized),
                "real_category": real_category,
                "all_real_categories": all_real_categories,
                "v2_category_candidate": classification["platform_category"],
                **classification,
                "review_priority": priority,
                "quality_flags": "；".join(quality_flags),
                "source_category_paths": all_real_categories,
                "source_catalog_ids": " | ".join(clean_text(value) for value in product.get("source_catalog_ids") or []),
                "price_min": price,
                "price_max": price,
                "stock_qty_snapshot": stock,
                "stock_status": "out_of_stock" if stock == 0 else "available",
                "sales_unit": clean_text(static.get("defaultUomName")),
                "source_status": clean_text(static.get("mshopProductStatusEnum")),
                "description_original": description_original,
                "description_candidate": description_candidate,
                "main_image_count": len(main_images),
                "detail_image_count": len(detail_images),
                "all_image_count": len(product_images),
                "images_pending_processing": noncompliant_count,
                "main_image_logical_path": clean_text(main_images[0].get("logical_relative_path")) if main_images else "",
                "manufacturer_name": SOURCE_PROVIDER_NAME,
                "manufacturer_status": "source_provider_label",
                "source_link": SOURCE_LINK_LABEL,
                "source_reference": f"{SOURCE_LINK_LABEL}；商品编码：{code}",
                "source_public_url": "",
            }
        )
    return result


def build_category_link_rows(
    links: list[dict[str, Any]], products_by_id: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for link in links:
        product_id = clean_text(link.get("product_id"))
        product = products_by_id.get(product_id, {})
        rows.append(
            {
                "source_product_id": product_id,
                "source_product_code": clean_text(product.get("source_product_code")),
                "normalized_name": clean_text(product.get("normalized_name")),
                "source_catalog_id": clean_text(link.get("product_catalog_id")),
                "source_catalog_name": clean_text(link.get("catalog_name")),
                "source_tree_path": clean_text(link.get("tree_path")).strip("^").replace("^", "/"),
                "real_category": clean_text(product.get("real_category")),
                "business_type": clean_text(product.get("business_type")),
                "v2_category_candidate": clean_text(product.get("v2_category_candidate")),
                "platform_category": clean_text(product.get("platform_category")),
                "classification_confidence": clean_text(product.get("classification_confidence")),
            }
        )
    return rows


def build_cluster_frame(category_links: list[dict[str, Any]]) -> pd.DataFrame:
    if not category_links:
        return pd.DataFrame(
            columns=[
                "source_category_l1",
                "source_category_l2",
                "business_type",
                "platform_category",
                "classification_confidence",
                "product_count",
                "representative_products",
            ]
        )
    frame = pd.DataFrame(category_links)
    parts = frame["source_tree_path"].str.split("/")
    frame["source_category_l1"] = parts.str[0].fillna("")
    frame["source_category_l2"] = parts.str[1].fillna("")
    grouped = (
        frame.groupby(
            [
                "source_category_l1",
                "source_category_l2",
                "business_type",
                "platform_category",
                "classification_confidence",
            ],
            dropna=False,
        )
        .agg(
            product_count=("source_product_code", "nunique"),
            representative_products=("normalized_name", lambda values: "；".join(dict.fromkeys(values).keys())[:500]),
        )
        .reset_index()
    )
    confidence_order = {"低": 0, "中": 1, "高": 2}
    grouped["_confidence"] = grouped["classification_confidence"].map(confidence_order).fillna(-1)
    grouped = grouped.sort_values(
        ["_confidence", "product_count", "source_category_l1", "source_category_l2"],
        ascending=[True, False, True, True],
    ).drop(columns=["_confidence"])
    return grouped


def quality_rows(
    products: list[dict[str, Any]], image_rows: list[dict[str, Any]], link_modes: Counter, batch_id: str
) -> list[dict[str, Any]]:
    business_counts = Counter(row["business_type"] for row in products)
    category_counts = Counter(row["platform_category"] or "未映射" for row in products)
    metrics = [
        ("capture_batch", batch_id, "本成果唯一输入批次"),
        ("unique_products", len(products), "按 source_product_code 唯一"),
        ("real_category_complete", sum(bool(row["real_category"]) for row in products), "真实来源主分类覆盖数量"),
        ("source_provider_complete", sum(bool(row["manufacturer_name"]) for row in products), "来源方标识覆盖数量"),
        ("auto_category_candidates", sum(row["review_status"] == "auto_candidate" for row in products), "高置信度分类候选"),
        ("classification_needs_review", sum(row["review_status"] == "needs_review" for row in products), "需要业务复核"),
        ("products_without_images", sum(row["all_image_count"] == 0 for row in products), "无任何图片关系"),
        ("products_without_source_description", sum(not row["description_original"] for row in products), "已生成描述候选但需抽检"),
        ("products_out_of_stock", sum(row["stock_status"] == "out_of_stock" for row in products), "采集时点库存为0"),
        ("image_relationships", len(image_rows), "商品与图片逻辑关系"),
        ("images_import_ready", sum(bool(row["import_image_ready"]) for row in image_rows), "同时满足下载、500KB和1000px限制"),
        ("images_pending_processing", sum(not bool(row["import_image_ready"]) for row in image_rows), "正式导入前需压缩或检查"),
        ("logical_images_hardlinked", link_modes.get("hardlink", 0), "不复制原图数据"),
        ("logical_images_copied", link_modes.get("copy", 0), "硬链接失败时的回退"),
    ]
    metrics.extend((f"business_type::{key}", value, "业务分流数量") for key, value in sorted(business_counts.items()))
    metrics.extend((f"platform_category::{key}", value, "平台分类候选数量") for key, value in sorted(category_counts.items()))
    return [{"metric": key, "value": value, "note": note} for key, value, note in metrics]


def field_description_rows() -> list[dict[str, str]]:
    return [
        {"field": "source_product_code", "meaning": "漫立方当前批次稳定商品编码，作为清洗业务键", "source": "structured/products.jsonl"},
        {"field": "source_product_id", "meaning": "当前接口商品ID，仅用于本批次接口追溯", "source": "structured/products.jsonl"},
        {"field": "business_type", "meaning": "整机、配件、礼品、非核心工具或待复核分流", "source": "新名称与新来源类目规则"},
        {"field": "real_category", "meaning": "从完整来源分类中选择的主分类，保证真实分类不因 v2 缺口丢失", "source": "structured/products.jsonl"},
        {"field": "all_real_categories", "meaning": "商品关联的全部真实来源分类路径", "source": "structured/products.jsonl"},
        {"field": "v2_category_candidate", "meaning": "游艺圈 v2 十类白名单兼容候选，空值不影响真实分类完整性", "source": "新名称与新来源类目规则"},
        {"field": "platform_category", "meaning": "保留的 v2 分类候选兼容字段", "source": "新名称与新来源类目规则"},
        {"field": "review_status", "meaning": "auto_candidate 或 needs_review", "source": "分类置信度"},
        {"field": "logical_relative_path", "meaning": "平台逻辑图片路径，不改变原始哈希文件", "source": "图片映射"},
        {"field": "import_image_ready", "meaning": "图片是否满足 500KB、1000px 和可读取要求", "source": "本地图片检测"},
        {"field": "manufacturer_name", "meaning": "按当前业务约定统一使用漫立方作为来源方标签，不代表已核验生产厂家主体", "source": "业务约定"},
        {"field": "source_link", "meaning": "来源入口标签：漫立方配件商城（微信小程序）", "source": "业务约定"},
        {"field": "source_reference", "meaning": "来源入口标签与稳定商品编码组合的逐商品追溯标识", "source": "业务约定与 source_product_code"},
        {"field": "source_public_url", "meaning": "暂无真实公开网页 URL，保持空值以避免伪造链接", "source": "缺失字段声明"},
    ]


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in worksheet.columns:
            letter = column[0].column_letter
            max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column[:200]), default=0)
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 48)
    workbook.save(path)


def build_clean_assets(batch_dir: Path, *, date_tag: str | None = None) -> dict[str, Any]:
    batch_dir = Path(batch_dir).resolve()
    structured = batch_dir / "structured"
    required = [
        structured / "products.jsonl",
        structured / "product_category_links.jsonl",
        structured / "static_details.jsonl",
        structured / "dynamic_details.jsonl",
        structured / "spu_details.jsonl",
        batch_dir / "discovered_image_urls.jsonl",
        batch_dir / "downloaded_image_manifest.jsonl",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("missing fresh capture inputs: " + ", ".join(missing))

    products = read_jsonl(structured / "products.jsonl")
    codes = [normalize_product_code(row.get("product_code")) for row in products]
    duplicates = [code for code, count in Counter(codes).items() if count > 1]
    if duplicates:
        raise ValueError(f"duplicate source product codes: {duplicates[:10]}")

    metadata_path = batch_dir / "batch_metadata.json"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8")) if metadata_path.exists() else {}
    batch_id = clean_text(metadata.get("batch_id")) or batch_dir.name
    product_code_by_id = {
        clean_text(row.get("product_id")): normalize_product_code(row.get("product_code")) for row in products
    }

    image_rows = build_image_rows(batch_dir)
    image_rows = assign_logical_image_names(image_rows, product_code_by_id)
    enrich_image_rows(batch_dir, image_rows)
    link_modes = create_logical_images(batch_dir, image_rows)

    product_rows = build_product_rows(
        products,
        read_jsonl(structured / "static_details.jsonl"),
        read_jsonl(structured / "dynamic_details.jsonl"),
        read_jsonl(structured / "spu_details.jsonl"),
        image_rows,
        batch_id,
    )
    products_by_id = {clean_text(row["source_product_id"]): row for row in product_rows}
    category_link_rows = build_category_link_rows(
        read_jsonl(structured / "product_category_links.jsonl"), products_by_id
    )
    cluster_frame = build_cluster_frame(category_link_rows)
    review_rows = [row for row in product_rows if row["review_priority"] in {"P0", "P1"}]

    output_dir = batch_dir / "cleaned"
    output_dir.mkdir(parents=True, exist_ok=True)
    tag = date_tag or date.today().strftime("%Y%m%d")
    workbook_path = output_dir / f"漫立方_新全量清洗主数据_{tag}.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(quality_rows(product_rows, image_rows, link_modes, batch_id)).to_excel(
            writer, sheet_name="清洗质量", index=False
        )
        pd.DataFrame(product_rows).to_excel(writer, sheet_name="商品清洗主表", index=False)
        cluster_frame.to_excel(writer, sheet_name="类目聚类", index=False)
        pd.DataFrame(category_link_rows).to_excel(writer, sheet_name="商品类目关系", index=False)
        pd.DataFrame(image_rows).to_excel(writer, sheet_name="图片映射", index=False)
        review_frame = pd.DataFrame(review_rows, columns=list(product_rows[0].keys()) if product_rows else [])
        review_frame.to_excel(writer, sheet_name="复核队列", index=False)
        pd.DataFrame(field_description_rows()).to_excel(writer, sheet_name="字段说明", index=False)
    style_workbook(workbook_path)

    return {
        "workbook": workbook_path,
        "product_count": len(product_rows),
        "category_link_count": len(category_link_rows),
        "image_relationship_count": len(image_rows),
        "logical_image_count": sum(link_modes.values()),
        "review_count": len(review_rows),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Clean a fresh Manlifang full capture into auditable XLSX assets")
    parser.add_argument("batch_dir", type=Path)
    parser.add_argument("--date-tag", default=None)
    args = parser.parse_args()
    result = build_clean_assets(args.batch_dir, date_tag=args.date_tag)
    print("clean_summary", " ".join(f"{key}={value}" for key, value in result.items()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
