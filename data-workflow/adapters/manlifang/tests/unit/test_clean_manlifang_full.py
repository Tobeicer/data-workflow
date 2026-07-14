from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


SOURCE_DIR = Path(__file__).resolve().parents[2] / "src"
if str(SOURCE_DIR) not in sys.path:
    sys.path.insert(0, str(SOURCE_DIR))

from clean_manlifang_full import (  # noqa: E402
    assign_logical_image_names,
    build_clean_assets,
    classify_product,
    choose_primary_source_category,
    logical_image_name,
)


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def test_logical_image_name_uses_stable_code_role_sequence_and_hash() -> None:
    name = logical_image_name(" C/01667 ", "primary", 2, "abcdef0123456789", ".jpeg")

    assert name == "MLF_C_01667_main_02_abcdef01.jpg"


def test_assign_logical_image_names_numbers_each_product_role_independently() -> None:
    rows = [
        {"product_id": "1", "image_role": "main", "sha256": "a" * 64, "local_file": "a.jpg"},
        {"product_id": "1", "image_role": "main", "sha256": "b" * 64, "local_file": "b.jpg"},
        {"product_id": "1", "image_role": "detail", "sha256": "c" * 64, "local_file": "c.png"},
        {"product_id": "2", "image_role": "main", "sha256": "d" * 64, "local_file": "d.jpg"},
    ]

    result = assign_logical_image_names(rows, {"1": "B100", "2": "B200"})

    assert [row["logical_file_name"] for row in result] == [
        "MLF_B100_main_01_aaaaaaaa.jpg",
        "MLF_B100_main_02_bbbbbbbb.jpg",
        "MLF_B100_detail_01_cccccccc.png",
        "MLF_B200_main_01_dddddddd.jpg",
    ]


def test_classify_product_keeps_non_platform_goods_out_of_v2_categories() -> None:
    coin = classify_product("漫立方投币器 MLF-168", ["投币器类^侧投式^"])
    machine = classify_product("儿童赛车游戏机整机", ["游戏机台^赛车摩托^"])
    tool = classify_product("OTS空压机750-30L", ["空压机 整机/配件^空压机（整机）^"])

    assert coin == {
        "business_type": "配件候选",
        "platform_category": "币器/投币器",
        "classification_confidence": "高",
        "classification_reason": "名称和来源类目均命中币器/投币器规则",
        "review_status": "auto_candidate",
    }
    assert machine["business_type"] == "整机产品候选"
    assert machine["platform_category"] == ""
    assert machine["review_status"] == "needs_review"
    assert tool["business_type"] == "非核心设备/工具候选"
    assert tool["platform_category"] == ""
    assert tool["review_status"] == "needs_review"


def test_classify_product_does_not_treat_machine_context_as_gift_or_tool() -> None:
    sensor = classify_product(
        "光眼：篮球机常开槽型光眼",
        ["机台套件^篮球机^", "电子器材^光眼计数^", "游戏机台^篮球机^光眼^"],
    )
    motor = classify_product(
        "电机：娃娃机迷你 左右2个螺丝孔",
        ["机台套件^精品娃娃机^", "电子器材^天车配件系列^", "游戏机台^娃娃机^电机^"],
    )
    ball = classify_product("篮球：棕色加厚5号", ["机台套件^篮球机^", "球类类目^篮球系列^"])

    assert sensor["business_type"] == "配件候选"
    assert sensor["platform_category"] == ""
    assert motor["business_type"] == "配件候选"
    assert motor["platform_category"] == "礼品机配件"
    assert ball["business_type"] == "礼品/玩法道具候选"
    assert ball["platform_category"] == "礼品"


def test_choose_primary_source_category_prefers_real_specialist_category() -> None:
    category = choose_primary_source_category(
        ["新品上新^", "机台套件^精品娃娃机^", "电子器材^天车配件系列^", "游戏机台^娃娃机^电机^"]
    )

    assert category == "电子器材/天车配件系列"
    assert choose_primary_source_category(["新品上新^"]) == "新品上新"


def test_build_clean_assets_uses_new_batch_and_creates_auditable_workbook(tmp_path: Path) -> None:
    structured = tmp_path / "structured"
    write_jsonl(
        structured / "products.jsonl",
        [
            {
                "product_id": 9001,
                "product_code": "B10001",
                "product_name": "漫立方投币器 MLF-168",
                "source_catalog_ids": [2],
                "source_tree_paths": ["投币器类^侧投式^"],
                "raw_listing": {"vRetailPrice": 12.5},
            }
        ],
    )
    write_jsonl(
        structured / "product_category_links.jsonl",
        [{"product_id": 9001, "product_catalog_id": 2, "catalog_name": "侧投式", "tree_path": "投币器类^侧投式^"}],
    )
    write_jsonl(
        structured / "static_details.jsonl",
        [
            {
                "productId": 9001,
                "name": "漫立方投币器 MLF-168",
                "code": "B10001",
                "displayDescription": "标准投币器",
                "defaultUomName": "个",
                "mshopProductStatusEnum": "ON_SHELF",
                "vRetailPriceForRetailUom": 12.5,
            }
        ],
    )
    write_jsonl(structured / "dynamic_details.jsonl", [{"productId": 9001, "price": 12.5, "retailPrice": 12.5, "availQty": 8}])
    write_jsonl(structured / "spu_details.jsonl", [{"productId": 9001, "fields": [{"name": "PRICE", "value": 12.5}]}])
    write_jsonl(
        structured / "categories.jsonl",
        [{"product_catalog_id": 2, "name": "侧投式", "tree_path": "投币器类^侧投式^", "is_leaf_node": True}],
    )

    image_path = tmp_path / "raw" / "images_downloaded" / "source.png"
    image_path.parent.mkdir(parents=True)
    Image.new("RGB", (20, 10), "white").save(image_path)
    content = image_path.read_bytes()
    digest = hashlib.sha256(content).hexdigest()
    image_url = "https://img.example.com/a.png"
    write_jsonl(
        tmp_path / "discovered_image_urls.jsonl",
        [{"product_id": 9001, "image_role": "main", "url": image_url, "json_path": "$.primaryImageList[0]", "source_endpoint": "static"}],
    )
    write_jsonl(
        tmp_path / "downloaded_image_manifest.jsonl",
        [{"url": image_url, "status": "downloaded", "content_type": "image/png", "bytes": len(content), "sha256": digest, "local_file": "raw/images_downloaded/source.png"}],
    )
    (tmp_path / "batch_metadata.json").write_text(json.dumps({"batch_id": "fresh_test"}), encoding="utf-8")

    result = build_clean_assets(tmp_path, date_tag="20260712")

    workbook = load_workbook(result["workbook"], read_only=True, data_only=True)
    assert workbook.sheetnames == ["清洗质量", "商品清洗主表", "类目聚类", "商品类目关系", "图片映射", "复核队列", "字段说明"]
    assert workbook["商品清洗主表"].max_row == 2
    assert workbook["图片映射"].max_row == 2
    assert workbook["复核队列"].max_row == 1
    product_sheet = workbook["商品清洗主表"]
    headers = [cell.value for cell in next(product_sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(product_sheet.iter_rows(min_row=2, max_row=2))]
    product = dict(zip(headers, values))
    assert product["real_category"] == "投币器类/侧投式"
    assert product["all_real_categories"] == "投币器类/侧投式"
    assert product["v2_category_candidate"] == "币器/投币器"
    assert product["manufacturer_name"] == "漫立方"
    assert product["source_link"] == "漫立方配件商城（微信小程序）"
    assert product["source_reference"] == "漫立方配件商城（微信小程序）；商品编码：B10001"
    assert product["source_public_url"] is None
    assert result["product_count"] == 1
    assert result["logical_image_count"] == 1
    logical_files = list((tmp_path / "cleaned" / "images" / "B10001").glob("*.png"))
    assert [path.name for path in logical_files] == [f"MLF_B10001_main_01_{digest[:8]}.png"]
    assert logical_files[0].read_bytes() == content
