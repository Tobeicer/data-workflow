from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

import pandas as pd


ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                value = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(value, dict):
                rows.append(value)
    return rows


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    value = ILLEGAL_XML_RE.sub("", value)
    return value if len(value) <= 32_000 else value[:32_000] + "...[truncated]"


def flatten_object(value: dict[str, Any], prefix: str = "", depth: int = 0) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, item in value.items():
        column = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(item, dict) and depth < 2:
            result.update(flatten_object(item, column, depth + 1))
        else:
            result[column] = excel_value(item)
    return result


def first_nonempty(*values: Any) -> Any:
    for value in values:
        if value not in (None, "", [], {}):
            return value
    return ""


def nested_id(value: Any) -> Any:
    return value.get("id", "") if isinstance(value, dict) else value


def as_joined(value: Any) -> str:
    if not isinstance(value, list):
        return excel_value(value)
    return " | ".join(str(item) for item in value if item not in (None, ""))


def spu_fields(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field in row.get("fields") or []:
        if not isinstance(field, dict):
            continue
        name = str(field.get("name", "")).strip()
        if name:
            result[name] = field.get("value", "")
    return result


def index_by(rows: Iterable[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = row.get(key)
        if value not in (None, ""):
            result[str(value)] = row
    return result


def build_image_rows(batch_dir: Path) -> list[dict[str, Any]]:
    discovered = read_jsonl(batch_dir / "discovered_image_urls.jsonl")
    downloaded = read_jsonl(batch_dir / "downloaded_image_manifest.jsonl")
    captured = read_jsonl(batch_dir / "captured_image_manifest.jsonl")
    download_by_url: dict[str, dict[str, Any]] = {}
    for row in captured + downloaded:
        url = str(row.get("url", ""))
        if not url:
            continue
        current = download_by_url.get(url)
        if current is None or row.get("status") == "downloaded":
            download_by_url[url] = row

    rows: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for row in discovered:
        url = str(row.get("url") or row.get("canonical_url") or row.get("original_url") or "")
        if not url:
            continue
        seen_urls.add(url)
        download = download_by_url.get(url, {})
        rows.append(
            {
                "product_id": row.get("product_id", ""),
                "image_role": row.get("image_role", ""),
                "source_endpoint": row.get("source_endpoint", ""),
                "json_path": row.get("json_path", ""),
                "original_url": row.get("original_url", url),
                "canonical_url": row.get("canonical_url", url),
                "url": url,
                "download_status": download.get("status", "captured" if download else "pending"),
                "status_code": download.get("status_code", ""),
                "content_type": download.get("content_type", ""),
                "bytes": download.get("bytes", ""),
                "sha256": download.get("sha256", ""),
                "local_file": download.get("local_file", ""),
                "download_error": download.get("error", ""),
            }
        )
    for url, download in download_by_url.items():
        if url in seen_urls:
            continue
        rows.append(
            {
                "product_id": "",
                "image_role": download.get("image_role", ""),
                "source_endpoint": "",
                "json_path": download.get("json_path", ""),
                "original_url": url,
                "canonical_url": url,
                "url": url,
                "download_status": download.get("status", "captured"),
                "status_code": download.get("status_code", ""),
                "content_type": download.get("content_type", ""),
                "bytes": download.get("bytes", ""),
                "sha256": download.get("sha256", ""),
                "local_file": download.get("local_file", ""),
                "download_error": download.get("error", ""),
            }
        )
    return rows


def category_rows_for_excel(categories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for category in categories:
        rows.append(
            {
                "product_catalog_id": category.get("product_catalog_id", ""),
                "parent_catalog_id": category.get("parent_catalog_id", ""),
                "name": category.get("name", ""),
                "tree_path": category.get("tree_path", ""),
                "depth": category.get("depth", ""),
                "is_leaf_node": category.get("is_leaf_node", ""),
                "is_hidden": category.get("is_hidden", ""),
                "sequence_num": category.get("sequence_num", ""),
                "product_catalog_type_id": category.get("product_catalog_type_id", ""),
                "image_url": category.get("image_url", ""),
                "raw_category_json": excel_value(category.get("raw_category", {})),
            }
        )
    return rows


def product_rows_for_excel(
    products: list[dict[str, Any]],
    static_rows: list[dict[str, Any]],
    dynamic_rows: list[dict[str, Any]],
    spu_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    static_by_id = index_by(static_rows, "productId")
    dynamic_by_id = index_by(dynamic_rows, "productId")
    spu_by_id = index_by(spu_rows, "productId")
    image_counts = Counter(str(row.get("product_id", "")) for row in image_rows if row.get("product_id") not in (None, ""))
    rows: list[dict[str, Any]] = []

    for product in products:
        product_id = product.get("product_id", "")
        key = str(product_id)
        raw_product = product.get("raw_product") or {}
        listing = product.get("raw_listing") or {}
        static = static_by_id.get(key, {})
        dynamic = dynamic_by_id.get(key, {})
        spu = spu_by_id.get(key, {})
        fields = spu_fields(spu)
        primary_images = first_nonempty(static.get("primaryImageList"), raw_product.get("primaryImageList"), [])
        detail_images = raw_product.get("detailImageList") or []
        rows.append(
            {
                "product_id": product_id,
                "listing_id": listing.get("id", ""),
                "product_code": first_nonempty(static.get("code"), product.get("product_code"), fields.get("CODE")),
                "product_name": first_nonempty(static.get("name"), product.get("product_name"), fields.get("NAME")),
                "source_catalog_ids": as_joined(product.get("source_catalog_ids", [])),
                "source_tree_paths": as_joined(product.get("source_tree_paths", [])),
                "is_multi_spec_enabled": first_nonempty(
                    static.get("isMultiSpecEnabled"),
                    spu.get("isMultiSpecEnabled"),
                    product.get("is_multi_spec_enabled"),
                ),
                "is_multi_spec_pregenerated": first_nonempty(
                    static.get("isMultiSpecPregenerated"), spu.get("isMultiSpecPregenerated")
                ),
                "is_product_sold_out": first_nonempty(static.get("isProductSoldOut"), spu.get("isProductSoldOut")),
                "mshop_product_status": static.get("mshopProductStatusEnum", ""),
                "product_type_id": nested_id(static.get("productTypeId", raw_product.get("productTypeId", ""))),
                "default_uom_id": static.get("defaultUomId", ""),
                "default_uom_name": static.get("defaultUomName", ""),
                "base_uom_id": nested_id(static.get("baseUomId", raw_product.get("baseUomId", ""))),
                "base_uom_name": first_nonempty(static.get("baseUomName"), (raw_product.get("baseUomId") or {}).get("uomName", "") if isinstance(raw_product.get("baseUomId"), dict) else ""),
                "retail_sales_uom_id": nested_id(static.get("retailSalesUomId", raw_product.get("retailSalesUomId", ""))),
                "retail_sales_uom_name": static.get("retailSalesUomName", ""),
                "listing_retail_price": listing.get("vRetailPrice", ""),
                "listing_member_price": listing.get("vMemberPrice", ""),
                "static_retail_price": static.get("vRetailPriceForRetailUom", ""),
                "dynamic_price": dynamic.get("price", ""),
                "dynamic_retail_price": dynamic.get("retailPrice", ""),
                "dynamic_avail_qty": dynamic.get("availQty", ""),
                "spu_price": fields.get("PRICE", ""),
                "spu_member_price": fields.get("MEMBER_PRICE", ""),
                "spu_retail_price": fields.get("RETAIL_PRICE", ""),
                "spu_avail_qty": fields.get("AVAIL_QTY", ""),
                "spu_month_sale_amount": fields.get("MONTH_SALE_AMOUNT", ""),
                "spu_spec": fields.get("SPEC", ""),
                "spu_spec_name": fields.get("SPEC_NAME", ""),
                "spu_spec_avail_qty": fields.get("SPEC_AVAIL_QTY", ""),
                "product_barcode": fields.get("PRODUCT_BARCODE", ""),
                "listing_sales_qty": listing.get("salesQty", ""),
                "listing_sales_unit_qty": listing.get("salesUnitQty", ""),
                "description": static.get("displayDescription", ""),
                "primary_image_count": len(primary_images) if isinstance(primary_images, list) else 0,
                "detail_image_count": len(detail_images) if isinstance(detail_images, list) else 0,
                "all_image_count": image_counts.get(key, 0),
                "primary_image_urls": as_joined(primary_images),
                "detail_image_urls": as_joined(detail_images),
                "product_catalog_ids_static": as_joined(static.get("productCatalogIdList", [])),
                "inventory_qty_dm_json": excel_value(raw_product.get("inventoryQtyDm", [])),
                "price_base_comp_json": excel_value(raw_product.get("priceBaseComp", [])),
                "raw_listing_json": excel_value(listing),
                "raw_product_json": excel_value(raw_product),
                "static_detail_json": excel_value(static),
                "dynamic_detail_json": excel_value(dynamic),
                "spu_detail_json": excel_value(spu),
            }
        )
    return rows


def detail_rows_for_excel(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [flatten_object(row) for row in rows]


def spu_rows_for_excel(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        fields = spu_fields(row)
        result.append(
            {
                "product_id": row.get("productId", ""),
                "uom_id": row.get("uomId", ""),
                "is_multi_spec_enabled": row.get("isMultiSpecEnabled", ""),
                "is_multi_spec_pregenerated": row.get("isMultiSpecPregenerated", ""),
                "is_product_sold_out": row.get("isProductSoldOut", ""),
                **{name.lower(): excel_value(value) for name, value in fields.items()},
                "raw_spu_json": excel_value(row),
            }
        )
    return result


def listing_rows_for_excel(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        result.append(
            {
                "product_id": row.get("product_id", ""),
                "listing_id": row.get("listing_id", ""),
                "product_code": row.get("product_code", ""),
                "product_name": row.get("product_name", ""),
                "source_catalog_id": row.get("source_catalog_id", ""),
                "source_catalog_name": row.get("source_catalog_name", ""),
                "source_tree_path": row.get("source_tree_path", ""),
                "page_first_result": row.get("page_first_result", ""),
                "page_count": row.get("page_count", ""),
                "raw_listing_json": excel_value(row.get("raw_listing", {})),
                "raw_product_json": excel_value(row.get("raw_product", {})),
            }
        )
    return result


def flow_rows_for_excel(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        result.append(
            {
                "flow_id": row.get("flow_id", ""),
                "captured_at": row.get("captured_at", ""),
                "endpoint": first_nonempty(row.get("endpoint"), row.get("path")),
                "request_key": row.get("request_key", ""),
                "request_body": excel_value(row.get("request_body", "")),
                "status_code": row.get("status_code", ""),
                "response_bytes": row.get("response_bytes", ""),
                "response_sha256": row.get("response_sha256", ""),
                "response_file": row.get("response_file", ""),
            }
        )
    return result


def quality_rows(
    batch_dir: Path,
    categories: list[dict[str, Any]],
    products: list[dict[str, Any]],
    links: list[dict[str, Any]],
    listings: list[dict[str, Any]],
    static_rows: list[dict[str, Any]],
    dynamic_rows: list[dict[str, Any]],
    spu_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
    flows: list[dict[str, Any]],
    errors: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    status_counts = Counter(row.get("status_code") for row in flows)
    endpoint_counts = Counter(str(first_nonempty(row.get("endpoint"), row.get("path"))) for row in flows)
    download_counts = Counter(str(row.get("download_status", "")) for row in image_rows)
    return [
        {"metric": "categories", "value": len(categories), "note": "实时类目节点总数"},
        {"metric": "leaf_categories", "value": sum(bool(row.get("is_leaf_node")) for row in categories), "note": "执行商品分页的叶子类目"},
        {"metric": "hidden_categories", "value": sum(bool(row.get("is_hidden")) for row in categories), "note": "接口标记为隐藏的类目"},
        {"metric": "unique_products", "value": len(products), "note": "按 product_id 去重"},
        {"metric": "product_category_links", "value": len(links), "note": "商品与类目多对多关系"},
        {"metric": "listing_records", "value": len(listings), "note": "全部类目分页中的商品记录"},
        {"metric": "static_details", "value": len(static_rows), "note": "静态详情成功结构化数量"},
        {"metric": "dynamic_details", "value": len(dynamic_rows), "note": "动态价格库存成功结构化数量"},
        {"metric": "spu_details", "value": len(spu_rows), "note": "SPU 字段商品数量"},
        {"metric": "multi_spec_products", "value": sum(bool(row.get("is_multi_spec_enabled")) for row in products), "note": "列表标记为多规格的商品"},
        {"metric": "unique_image_rows", "value": len(image_rows), "note": "按商品和原始图片路径发现的图片关系"},
        {"metric": "downloaded_images", "value": download_counts.get("downloaded", 0) + download_counts.get("captured", 0), "note": "已有本地文件或代理捕获"},
        {"metric": "failed_image_downloads", "value": download_counts.get("failed", 0), "note": "需要重试的图片"},
        {"metric": "api_flows", "value": len(flows), "note": "原始响应索引数量"},
        {"metric": "http_200_flows", "value": status_counts.get(200, 0), "note": "HTTP 200 响应"},
        {"metric": "non_200_flows", "value": len(flows) - status_counts.get(200, 0), "note": "需要复核的非 200 响应"},
        {"metric": "collector_errors", "value": len(errors), "note": "含运行中断等采集器错误记录"},
        {"metric": "endpoint_counts", "value": excel_value(dict(endpoint_counts)), "note": "各接口响应数量"},
        {"metric": "batch_dir", "value": str(batch_dir), "note": "原始批次目录"},
    ]


def frame(rows: list[dict[str, Any]], columns: list[str] | None = None) -> pd.DataFrame:
    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=columns or ["暂无数据"])


def autofit(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    workbook = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for cells in sheet.columns:
            max_length = 0
            for cell in cells[:300]:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[cells[0].column_letter].width = min(max(max_length + 2, 10), 48)
    workbook.save(path)


def build_workbook(batch_dir: Path, output: Path) -> Path:
    batch_dir = batch_dir.resolve()
    output = output.resolve()
    structured = batch_dir / "structured"
    categories = read_jsonl(structured / "categories.jsonl")
    products = read_jsonl(structured / "products.jsonl")
    links = read_jsonl(structured / "product_category_links.jsonl")
    listings = read_jsonl(structured / "listing_records.jsonl")
    static_rows = read_jsonl(structured / "static_details.jsonl")
    dynamic_rows = read_jsonl(structured / "dynamic_details.jsonl")
    spu_rows = read_jsonl(structured / "spu_details.jsonl")
    image_rows = build_image_rows(batch_dir)
    flows = read_jsonl(batch_dir / "api_flows.jsonl")
    errors = read_jsonl(batch_dir / "collector_errors.jsonl") + read_jsonl(batch_dir / "capture_errors.jsonl")

    sheets = {
        "采集质量": frame(quality_rows(batch_dir, categories, products, links, listings, static_rows, dynamic_rows, spu_rows, image_rows, flows, errors)),
        "类目树": frame(category_rows_for_excel(categories)),
        "商品主表": frame(product_rows_for_excel(products, static_rows, dynamic_rows, spu_rows, image_rows)),
        "商品类目关系": frame(links),
        "商品列表记录": frame(listing_rows_for_excel(listings)),
        "静态详情": frame(detail_rows_for_excel(static_rows)),
        "动态价格库存": frame(detail_rows_for_excel(dynamic_rows)),
        "SPU规格字段": frame(spu_rows_for_excel(spu_rows)),
        "图片清单": frame(image_rows),
        "接口响应索引": frame(flow_rows_for_excel(flows)),
        "异常记录": frame(errors),
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)
    autofit(output)
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the comprehensive structured Manlifang XLSX workbook")
    parser.add_argument("batch_dir", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    batch_dir = args.batch_dir.resolve()
    output = args.output or batch_dir / f"漫立方_原始全量商品数据_{batch_dir.name}.xlsx"
    result = build_workbook(batch_dir, output)
    print(f"workbook={result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
