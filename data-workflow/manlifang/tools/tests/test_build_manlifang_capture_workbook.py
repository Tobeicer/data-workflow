from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


MANLIFANG_DIR = Path(__file__).resolve().parents[1]
if str(MANLIFANG_DIR) not in sys.path:
    sys.path.insert(0, str(MANLIFANG_DIR))

from build_manlifang_capture_workbook import build_workbook  # noqa: E402


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def test_build_workbook_creates_exact_structured_sheets(tmp_path: Path) -> None:
    structured = tmp_path / "structured"
    write_jsonl(
        structured / "categories.jsonl",
        [
            {
                "product_catalog_id": 1,
                "parent_catalog_id": "",
                "name": "根类目",
                "tree_path": "根类目^",
                "is_leaf_node": False,
                "is_hidden": False,
                "depth": 0,
                "raw_category": {"productCatalogId": 1},
            },
            {
                "product_catalog_id": 2,
                "parent_catalog_id": 1,
                "name": "叶子类目",
                "tree_path": "根类目^叶子类目^",
                "is_leaf_node": True,
                "is_hidden": False,
                "depth": 1,
                "raw_category": {"productCatalogId": 2},
            },
        ],
    )
    write_jsonl(
        structured / "products.jsonl",
        [
            {
                "product_id": 9001,
                "product_code": "B10001",
                "product_name": "测试商品",
                "is_multi_spec_enabled": False,
                "source_catalog_ids": [2],
                "source_tree_paths": ["根类目^叶子类目^"],
                "raw_product": {
                    "id": 9001,
                    "baseUomId": {"id": 501, "uomName": "个"},
                    "primaryImageList": ["https://img.example.com/a.jpg"],
                    "detailImageList": ["https://img.example.com/d.jpg"],
                },
                "raw_listing": {"id": 7001, "vRetailPrice": 12.5, "salesQty": 3},
            }
        ],
    )
    write_jsonl(
        structured / "product_category_links.jsonl",
        [{"product_id": 9001, "product_catalog_id": 2, "catalog_name": "叶子类目", "tree_path": "根类目^叶子类目^"}],
    )
    write_jsonl(
        structured / "listing_records.jsonl",
        [{"product_id": 9001, "listing_id": 7001, "source_catalog_id": 2, "raw_listing": {"id": 7001}}],
    )
    write_jsonl(
        structured / "static_details.jsonl",
        [
            {
                "productId": 9001,
                "name": "测试商品详情",
                "code": "B10001",
                "displayDescription": "完整描述",
                "primaryImageList": ["https://img.example.com/a.jpg", "https://img.example.com/b.jpg"],
                "vRetailPriceForRetailUom": 13.0,
                "defaultUomName": "个",
            }
        ],
    )
    write_jsonl(
        structured / "dynamic_details.jsonl",
        [{"productId": 9001, "price": 11.0, "retailPrice": 13.0, "availQty": 8}],
    )
    write_jsonl(
        structured / "spu_details.jsonl",
        [
            {
                "productId": 9001,
                "uomId": 501,
                "isMultiSpecEnabled": False,
                "fields": [
                    {"name": "PRICE", "value": 11.0},
                    {"name": "AVAIL_QTY", "value": 8},
                    {"name": "SPEC_NAME", "value": ""},
                ],
            }
        ],
    )
    write_jsonl(
        tmp_path / "discovered_image_urls.jsonl",
        [
            {
                "product_id": 9001,
                "url": "https://img.example.com/a.jpg",
                "canonical_url": "https://img.example.com/a.jpg",
                "original_url": "https://img.example.com/a.jpg?resize=200",
                "image_role": "main",
                "source_endpoint": "static_detail",
                "json_path": "$.primaryImageList[0]",
            }
        ],
    )
    write_jsonl(
        tmp_path / "downloaded_image_manifest.jsonl",
        [
            {
                "url": "https://img.example.com/a.jpg",
                "status": "downloaded",
                "sha256": "abc",
                "local_file": "raw/images_downloaded/abc.jpg",
                "bytes": 123,
                "content_type": "image/jpeg",
            }
        ],
    )
    write_jsonl(
        tmp_path / "api_flows.jsonl",
        [
            {
                "flow_id": "flow-1",
                "captured_at": "2026-07-10T00:00:00+08:00",
                "endpoint": "static_detail",
                "request_key": "9001",
                "request_body": {"productId": 9001},
                "status_code": 200,
                "response_file": "raw/responses/static_detail/9001.json",
            }
        ],
    )

    output = tmp_path / "result.xlsx"
    build_workbook(tmp_path, output)

    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "采集质量",
        "类目树",
        "商品主表",
        "商品类目关系",
        "商品列表记录",
        "静态详情",
        "动态价格库存",
        "SPU规格字段",
        "图片清单",
        "接口响应索引",
        "异常记录",
    ]
    assert workbook["类目树"].max_row == 3
    assert workbook["商品主表"].max_row == 2
    headers = [cell.value for cell in next(workbook["商品主表"].iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(workbook["商品主表"].iter_rows(min_row=2, max_row=2))]
    row = dict(zip(headers, values))
    assert row["product_id"] == 9001
    assert row["product_name"] == "测试商品详情"
    assert row["dynamic_avail_qty"] == 8
    assert row["spu_price"] == 11
    assert row["all_image_count"] == 1
    image_headers = [cell.value for cell in next(workbook["图片清单"].iter_rows(min_row=1, max_row=1))]
    image_values = [cell.value for cell in next(workbook["图片清单"].iter_rows(min_row=2, max_row=2))]
    image_row = dict(zip(image_headers, image_values))
    assert image_row["download_status"] == "downloaded"
    assert image_row["local_file"] == "raw/images_downloaded/abc.jpg"
