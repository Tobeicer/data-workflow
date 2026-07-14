from __future__ import annotations

import hashlib
import json
import sys
import uuid
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image


MANLIFANG_DIR = Path(__file__).resolve().parents[1]
if str(MANLIFANG_DIR) not in sys.path:
    sys.path.insert(0, str(MANLIFANG_DIR))

from build_manlifang_delivery_package import (  # noqa: E402
    DELIVERY_SHEETS,
    build_delivery_package,
    normalize_delivery_image,
    prepare_delivery_image,
    stable_uid,
)


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def test_stable_uid_is_deterministic_and_source_namespaced() -> None:
    first = stable_uid("source_product", "manlifang", "B10001")
    second = stable_uid("source_product", "manlifang", "B10001")
    other_source = stable_uid("source_product", "1688", "B10001")

    assert first == second
    assert first != other_source
    assert str(uuid.UUID(first)) == first


def test_normalize_delivery_image_enforces_size_and_dimensions(tmp_path: Path) -> None:
    source = tmp_path / "large.png"
    target = tmp_path / "out.jpg"
    Image.effect_noise((2200, 1600), 120).convert("RGB").save(source)

    result = normalize_delivery_image(source, target)

    assert result["bytes"] <= 500 * 1024
    assert result["width"] <= 1000
    assert result["height"] <= 1000
    assert result["content_type"] == "image/jpeg"
    with Image.open(target) as image:
        assert image.format == "JPEG"


def test_prepare_delivery_image_normalizes_mislabeled_non_jpeg(tmp_path: Path) -> None:
    source = tmp_path / "source.gif"
    target = tmp_path / "out.jpg"
    Image.new("RGB", (100, 100), "white").save(source, format="GIF")

    result = prepare_delivery_image(
        source,
        target,
        {"content_type": "image/jpeg", "bytes": source.stat().st_size, "width_px": 100, "height_px": 100},
    )

    assert result["processing"] == "normalized"
    with Image.open(target) as image:
        assert image.format == "JPEG"


def test_build_delivery_package_creates_excel_tables_and_deduplicated_images(tmp_path: Path) -> None:
    batch = tmp_path / "batch"
    structured = batch / "structured"
    clean_workbook = batch / "cleaned" / "clean.xlsx"
    output_dir = tmp_path / "delivery" / "漫立方_全量数据"
    source_image = batch / "raw" / "images_downloaded" / "source.png"
    source_image.parent.mkdir(parents=True)
    Image.new("RGB", (1200, 800), "white").save(source_image)
    content = source_image.read_bytes()
    digest = hashlib.sha256(content).hexdigest()

    products = pd.DataFrame(
        [
            {
                "source_system": "manlifang",
                "capture_batch": "test_batch",
                "source_product_id": "9001",
                "source_product_code": "B10001",
                "original_name": "测试投币器",
                "normalized_name": "测试投币器",
                "model_candidate": "MLF-168",
                "real_category": "投币器类/侧投式",
                "all_real_categories": "投币器类/侧投式",
                "v2_category_candidate": "币器/投币器",
                "business_type": "配件候选",
                "classification_confidence": "高",
                "classification_reason": "测试",
                "review_status": "auto_candidate",
                "review_priority": "P2",
                "quality_flags": "",
                "price_min": 12.5,
                "price_max": 12.5,
                "stock_qty_snapshot": 8,
                "stock_status": "available",
                "sales_unit": "个",
                "source_status": "ON_SHELF",
                "description_original": "标准投币器",
                "description_candidate": "标准投币器",
                "manufacturer_name": "漫立方",
                "source_link": "漫立方配件商城（微信小程序）",
                "source_reference": "漫立方配件商城（微信小程序）；商品编码：B10001",
            },
            {
                "source_system": "manlifang",
                "capture_batch": "test_batch",
                "source_product_id": "9002",
                "source_product_code": "B10002",
                "original_name": "无图按钮",
                "normalized_name": "无图按钮",
                "model_candidate": "",
                "real_category": "按钮类目/方形按钮",
                "all_real_categories": "按钮类目/方形按钮",
                "v2_category_candidate": "摇杆按钮",
                "business_type": "配件候选",
                "classification_confidence": "高",
                "classification_reason": "测试",
                "review_status": "auto_candidate",
                "review_priority": "P1",
                "quality_flags": "缺少图片",
                "price_min": 3,
                "price_max": 3,
                "stock_qty_snapshot": 0,
                "stock_status": "out_of_stock",
                "sales_unit": "个",
                "source_status": "ON_SHELF",
                "description_original": "",
                "description_candidate": "无图按钮",
                "manufacturer_name": "漫立方",
                "source_link": "漫立方配件商城（微信小程序）",
                "source_reference": "漫立方配件商城（微信小程序）；商品编码：B10002",
            },
        ]
    )
    category_links = pd.DataFrame(
        [
            {"source_product_id": "9001", "source_product_code": "B10001", "source_catalog_id": "10", "source_tree_path": "投币器类/侧投式"},
            {"source_product_id": "9002", "source_product_code": "B10002", "source_catalog_id": "11", "source_tree_path": "按钮类目/方形按钮"},
        ]
    )
    image_mapping = pd.DataFrame(
        [
            {
                "product_id": "9001",
                "product_code": "B10001",
                "image_role": "main",
                "image_sequence": 1,
                "original_url": "https://img.example.com/a.png",
                "source_endpoint": "static",
                "json_path": "$.primary[0]",
                "sha256": digest,
                "local_file": "raw/images_downloaded/source.png",
                "content_type": "image/png",
                "bytes": len(content),
            }
        ]
    )
    clean_workbook.parent.mkdir(parents=True)
    with pd.ExcelWriter(clean_workbook, engine="openpyxl") as writer:
        products.to_excel(writer, sheet_name="商品清洗主表", index=False)
        category_links.to_excel(writer, sheet_name="商品类目关系", index=False)
        image_mapping.to_excel(writer, sheet_name="图片映射", index=False)
        products.iloc[1:].to_excel(writer, sheet_name="复核队列", index=False)

    write_jsonl(
        structured / "categories.jsonl",
        [
            {"product_catalog_id": 10, "parent_catalog_id": "", "name": "侧投式", "tree_path": "投币器类^侧投式^", "depth": 1, "is_leaf_node": True},
            {"product_catalog_id": 11, "parent_catalog_id": "", "name": "方形按钮", "tree_path": "按钮类目^方形按钮^", "depth": 1, "is_leaf_node": True},
        ],
    )
    (batch / "batch_metadata.json").write_text(
        json.dumps({"batch_id": "test_batch", "source": "fresh"}), encoding="utf-8"
    )

    result = build_delivery_package(batch, clean_workbook, output_dir)

    workbook_path = output_dir / "漫立方_全量数据.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert workbook.sheetnames == DELIVERY_SHEETS
    assert workbook["v2导入"].max_row == 3
    assert workbook["来源商品"].max_row == 3
    assert workbook["来源类目"].max_row == 3
    assert workbook["商品类目关系"].max_row == 3
    assert workbook["图片"].max_row == 3
    assert workbook["商品图片关系"].max_row == 3
    assert workbook["价格库存"].max_row == 3
    assert all(cell.value not in (None, "") for row in workbook["v2导入"].iter_rows(min_row=2) for cell in row)

    images = sorted((output_dir / "images").glob("*.jpg"))
    assert len(images) == 2
    assert any(path.name == "MLFIMG_NO_IMAGE.jpg" for path in images)
    assert result["product_count"] == 2
    assert result["unique_image_count"] == 2
    assert result["product_image_relation_count"] == 2
