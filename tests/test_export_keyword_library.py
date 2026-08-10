import json
import sqlite3
import sys
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "tools"))

import export_keyword_library as ex  # noqa: E402


FIXTURE = {
    "version": "1.0.0",
    "source": "1688",
    "categories": [
        {
            "category_code": "A01",
            "category_name": "礼品抓取、售卖",
            "concepts": [
                {
                    "standard_name": "娃娃机",
                    "aliases": ["夹娃娃机", "抓娃机", "夹物机"],
                    "platforms": {
                        "1688": ["夹娃娃机"],
                        "taobao": ["抓娃机", "夹物机"],
                    },
                    "source": "taxonomy",
                    "status": "pending",
                },
                {
                    "standard_name": "剪刀机",
                    "aliases": ["切绳机"],
                    "platforms": {"1688": ["切绳机"]},
                    "source": "taxonomy",
                    "status": "active",
                },
            ],
        }
    ],
    "candidate_pool": [
        {
            "term": "魔术师",
            "frequency": 3,
            "sample_titles": ["商用魔术师游戏机"],
            "source": "title_mining",
            "status": "pending",
        }
    ],
}


def test_export_xlsx_builds_summary_sheet(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "kw.xlsx"
    ex.export_xlsx(FIXTURE, out)
    assert out.exists()
    wb = load_workbook(out, read_only=True)
    assert "全平台词库总表" in wb.sheetnames
    assert "候选池审校" in wb.sheetnames
    ws = wb["全平台词库总表"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # 表头 + 2 概念
    assert rows[1][0] == "A01"
    assert rows[1][2] == "娃娃机"
    assert "夹娃娃机" in rows[1][3]  # 全平台合并列
    assert "抓娃机" in rows[1][3]
    assert rows[1][4] == "taxonomy"  # 来源列
    ws2 = wb["候选池审校"]
    assert len(list(ws2.iter_rows(values_only=True))) == 2


def test_export_sqlite_creates_tables_and_data(tmp_path: Path) -> None:
    out = tmp_path / "kw.sqlite"
    ex.export_sqlite(FIXTURE, out)
    conn = sqlite3.connect(out)
    cur = conn.cursor()
    assert cur.execute("SELECT COUNT(*) FROM keyword_concept").fetchone()[0] == 2
    assert cur.execute("SELECT COUNT(*) FROM keyword_alias").fetchone()[0] >= 4
    assert cur.execute("SELECT COUNT(*) FROM keyword_candidate").fetchone()[0] == 1
    rows = cur.execute(
        """
        SELECT DISTINCT a.alias, a.platform FROM keyword_alias a
        JOIN keyword_concept c ON c.id = a.concept_id
        WHERE c.standard_name = '娃娃机' ORDER BY a.platform
        """
    ).fetchall()
    assert ("夹娃娃机", "1688") in rows
    assert ("抓娃机", "taobao") in rows
    assert ("夹物机", "all") in rows
    conn.close()


def test_main_end_to_end(tmp_path: Path, monkeypatch) -> None:
    lib = tmp_path / "keywords_all_platforms.json"
    lib.write_text(json.dumps(FIXTURE, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(
        "sys.argv",
        ["export_keyword_library.py", "--library", str(lib)],
    )
    ex.main()
    assert lib.with_suffix(".xlsx").exists()
    assert lib.with_suffix(".sqlite").exists()
