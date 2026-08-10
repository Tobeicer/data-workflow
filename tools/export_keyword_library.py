"""把全平台关键词库导出为 Excel 总表与 SQLite 数据库。

输入：`deliveries/keywords/keywords_all_platforms.json`（build_keyword_library 生成）。
输出：
- `deliveries/keywords/keywords_all_platforms.xlsx`：人工总表（概念一行，同义词合并列，精简 6 列）；
- `deliveries/keywords/keywords_all_platforms.sqlite`：数据库（keyword_concept / keyword_alias / keyword_candidate）。
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from pathlib import Path


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LIBRARY = REPO_ROOT / "deliveries" / "keywords" / "keywords_all_platforms.json"


def export_xlsx(config: dict, output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "全平台词库总表"
    cols = ["分类码", "分类名", "标准名", "同义词(差异明显别称)", "来源", "状态"]
    ws.append(cols)
    for c in cols:
        cell = ws.cell(row=1, column=cols.index(c) + 1)
        cell.fill = hf
        cell.font = hfont
    for cat in config["categories"]:
        for concept in cat["concepts"]:
            row = [
                cat["category_code"],
                cat["category_name"],
                concept["standard_name"],
                "、".join(concept["aliases"]),
                concept["source"],
                concept["status"],
            ]
            ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for i, w in enumerate([8, 22, 20, 55, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("候选池审校")
    cols2 = ["关键词", "频次", "样例标题(前3)", "来源", "状态"]
    ws2.append(cols2)
    for c in cols2:
        cell = ws2.cell(row=1, column=cols2.index(c) + 1)
        cell.fill = hf
        cell.font = hfont
    for cand in config.get("candidate_pool", []):
        ws2.append(
            [
                cand["term"],
                cand["frequency"],
                "\n".join(cand.get("sample_titles", [])[:3]),
                cand["source"],
                cand["status"],
            ]
        )
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:E{ws2.max_row}"
    for i, w in enumerate([24, 8, 60, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def export_sqlite(config: dict, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        output.unlink()
    conn = sqlite3.connect(output)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE keyword_concept (
            id INTEGER PRIMARY KEY,
            category_code TEXT NOT NULL,
            category_name TEXT,
            standard_name TEXT NOT NULL,
            source TEXT,
            status TEXT,
            UNIQUE(category_code, standard_name)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE keyword_alias (
            alias_id INTEGER PRIMARY KEY,
            concept_id INTEGER NOT NULL REFERENCES keyword_concept(id),
            alias TEXT NOT NULL,
            platform TEXT NOT NULL,
            UNIQUE(concept_id, alias, platform)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE keyword_candidate (
            id INTEGER PRIMARY KEY,
            term TEXT NOT NULL,
            frequency INTEGER,
            sample_titles TEXT,
            source TEXT,
            status TEXT
        )
        """
    )
    for cat in config["categories"]:
        for concept in cat["concepts"]:
            cur.execute(
                "INSERT INTO keyword_concept (category_code, category_name, standard_name, source, status) VALUES (?,?,?,?,?)",
                (
                    cat["category_code"],
                    cat["category_name"],
                    concept["standard_name"],
                    concept.get("source"),
                    concept.get("status"),
                ),
            )
            concept_id = cur.lastrowid
            for alias in concept.get("aliases", []):
                cur.execute(
                    "INSERT INTO keyword_alias (concept_id, alias, platform) VALUES (?,?,?)",
                    (concept_id, alias, "all"),
                )
            for platform, terms in concept.get("platforms", {}).items():
                for alias in terms:
                    cur.execute(
                        "INSERT OR IGNORE INTO keyword_alias (concept_id, alias, platform) VALUES (?,?,?)",
                        (concept_id, alias, platform),
                    )
    for cand in config.get("candidate_pool", []):
        cur.execute(
            "INSERT INTO keyword_candidate (term, frequency, sample_titles, source, status) VALUES (?,?,?,?,?)",
            (
                cand["term"],
                cand.get("frequency"),
                "\n".join(cand.get("sample_titles", [])[:3]),
                cand.get("source"),
                cand.get("status"),
            ),
        )
    conn.commit()
    conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="导出关键词库总表与数据库")
    parser.add_argument("--library", default=str(DEFAULT_LIBRARY), help="全平台词库 JSON")
    args = parser.parse_args()

    library_path = Path(args.library)
    config = json.loads(library_path.read_text(encoding="utf-8"))
    xlsx = library_path.with_suffix(".xlsx")
    sqlite = library_path.with_suffix(".sqlite")
    export_xlsx(config, xlsx)
    export_sqlite(config, sqlite)
    total = sum(len(c["concepts"]) for c in config["categories"])
    aliases = sum(
        1 + len(c["aliases"]) for cat in config["categories"] for c in cat["concepts"]
    )
    print(f"concepts={total} search_terms={aliases}")
    print(f"xlsx={xlsx}")
    print(f"sqlite={sqlite}")


if __name__ == "__main__":
    main()
